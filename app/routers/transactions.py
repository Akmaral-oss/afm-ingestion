"""
Transactions API:
- GET /api/v1/transactions
- POST /api/v1/transactions/import-statement
"""

import csv
import hashlib
import json
import os
from datetime import datetime
from io import BytesIO, StringIO
from tempfile import NamedTemporaryFile
from typing import Optional
import re
from uuid import uuid4

import httpx
from fastapi.concurrency import run_in_threadpool
from fastapi import APIRouter, Depends, Query, File, UploadFile, Header, HTTPException, status, Form
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
from sqlalchemy import select, func, and_, or_, case, text, cast, Date
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..database import get_db, async_session, async_engine
from ..ingestion.fraud import ImportedTransactionSample, detect_import_fraud_warnings
from ..ingestion.pipeline import IngestionPipeline
from ..models import EsfRecord, Transaction, TransactionUploadMeta
from ..project_context import ProjectContext, get_current_project_context
from ..schemas import (
    TransactionListResponse,
    TransactionOut,
    EsfCounterpartyOut,
    EsfListResponse,
    EsfRecordOut,
    EsfSheetResponse,
    EsfSummaryOut,
    CounterpartyOut,
    PaginationOut,
    SummaryOut,
    TransactionImportResponse,
)
from ..security import decode_access_token


router = APIRouter(prefix="/transactions", tags=["Transactions"])

PARSER_BANK = "bank_parser"
PARSER_KASPI = "kaspi"
PARSER_KASPI_LEGACY = "kaspi_parser"
PARSER_HALYK = "halyk_parser"
PARSER_SMART = "smart_parser"
PARSER_TRANSACTIONS_CORE = "transactions_core"
PARSER_ESF = "esf"
_KZT_CONVERSION_RATES = {
    "KZT": 1.0,
    "USD": 475.50,
    "EUR": 520.30,
    "RUB": 5.20,
    "CNY": 66.00,
}


def _parse_date(date_str: str) -> datetime:
    return datetime.strptime(date_str, "%d.%m.%Y")


def _effective_transaction_dt(tx: Transaction) -> Optional[datetime]:
    if tx.date is not None:
        return tx.date
    if tx.operation_date is not None:
        return datetime.combine(tx.operation_date, datetime.min.time())
    return None


def _format_transaction_dt(tx: Transaction) -> str:
    dt = _effective_transaction_dt(tx)
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""


def _parse_operation_datetime(raw: object) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw

    s = str(raw).strip()
    if not s:
        return None

    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _decode_text_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "utf-16le", "utf-16be", "cp1251", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _to_import_fraud_sample(tx: Transaction) -> ImportedTransactionSample:
    return ImportedTransactionSample(
        tx_id=str(tx.id),
        operation_date=tx.operation_date,
        operation_ts=tx.date,
        amount_kzt=float(tx.amount_tenge or 0),
        credit=float(tx.credit or 0),
        debit=float(tx.debit or 0),
        direction=str(tx.direction or ""),
        purpose_text=str(tx.purpose or ""),
        operation_type_raw=str(tx.operation_type or ""),
        transaction_category=str(tx.transaction_category or ""),
        payer_name=str(tx.sender_name or ""),
        payer_iin_bin=str(tx.sender_iin_bin or ""),
        payer_account=str(tx.sender_account or ""),
        receiver_name=str(tx.recipient_name or ""),
        receiver_iin_bin=str(tx.recipient_iin_bin or ""),
        receiver_account=str(tx.recipient_account or ""),
    )


def _serialize_import_fraud_warnings(transactions: list[Transaction]) -> list[dict]:
    warnings = detect_import_fraud_warnings(_to_import_fraud_sample(tx) for tx in transactions)
    return [
        {
            "code": warning.code,
            "title": warning.title,
            "severity": warning.severity,
            "summary": warning.summary,
            "articles": list(warning.articles),
            "indicators": [
                {"label": indicator.label, "value": indicator.value}
                for indicator in warning.indicators
            ],
            "counterparties": [
                {
                    "role": counterparty.role,
                    "name": counterparty.name,
                    "identifier": counterparty.identifier,
                    "transaction_count": counterparty.transaction_count,
                    "turnover": counterparty.turnover,
                    "articles": list(counterparty.articles),
                    "graph_iin_bin": counterparty.graph_iin_bin or None,
                }
                for counterparty in warning.counterparties
            ],
            "sample_transactions": [
                {
                    "tx_id": tx.tx_id,
                    "happened_at": tx.happened_at,
                    "direction": tx.direction,
                    "amount": tx.amount,
                    "counterparty": tx.counterparty,
                    "purpose": tx.purpose,
                    "sender_name": tx.sender_name,
                    "sender_iin_bin": tx.sender_iin_bin,
                    "sender_account": tx.sender_account,
                    "recipient_name": tx.recipient_name,
                    "recipient_iin_bin": tx.recipient_iin_bin,
                    "recipient_account": tx.recipient_account,
                }
                for tx in warning.sample_transactions
            ],
        }
        for warning in warnings
    ]


async def _build_import_fraud_warnings_for_file(
    file_id: Optional[str],
    project_id: Optional[str],
    db: Optional[AsyncSession] = None,
) -> list[dict]:
    if not file_id or not project_id:
        return []

    async def _load(session: AsyncSession) -> list[dict]:
        result = await session.execute(
            select(Transaction)
            .where(
                Transaction.file_id == file_id,
                Transaction.project_id == project_id,
            )
            .order_by(Transaction.operation_date.asc(), Transaction.date.asc())
        )
        transactions = result.scalars().all()
        if not transactions:
            return []
        return _serialize_import_fraud_warnings(transactions)

    if db is not None:
        return await _load(db)

    async with async_session() as session:
        return await _load(session)


_REF_NO_PATTERN = re.compile(r"№\s*([A-Za-zА-Яа-я0-9/_-]{3,})", re.IGNORECASE)
_REF_ESF_PATTERN = re.compile(r"(ESF[-A-Za-z0-9_]+)", re.IGNORECASE)


def _format_money_label(value: float) -> str:
    return f"{float(value or 0):,.0f} KZT".replace(",", " ")


def _normalize_ref_token(value: object) -> str:
    token = str(value or "").strip().upper()
    token = token.replace('"', "").replace("'", "")
    token = re.sub(r"\s+", "", token)
    return token


def _extract_reference_tokens(*values: object) -> set[str]:
    tokens: set[str] = set()
    for value in values:
        text = _fix_mojibake(value or "").strip()
        if not text:
            continue
        for match in _REF_NO_PATTERN.findall(text):
            token = _normalize_ref_token(match)
            if token:
                tokens.add(token)
        for match in _REF_ESF_PATTERN.findall(text):
            token = _normalize_ref_token(match)
            if token:
                tokens.add(token)
    return tokens


def _tx_amount_value(tx: Transaction) -> float:
    return max(float(tx.credit or 0), float(tx.debit or 0), float(tx.amount_tenge or 0))


def _serialize_tx_as_warning_sample(tx: Transaction) -> dict:
    return {
        "tx_id": str(tx.id),
        "happened_at": _format_transaction_dt(tx),
        "direction": "Входящая" if float(tx.credit or 0) > 0 else "Исходящая",
        "amount": _format_money_label(_tx_amount_value(tx)),
        "counterparty": _fix_mojibake(
            (tx.sender_name if float(tx.credit or 0) > 0 else tx.recipient_name)
            or tx.sender_account
            or tx.recipient_account
            or ""
        ) or "Не указан",
        "purpose": _fix_mojibake(tx.purpose or tx.raw_note or tx.operation_type or "")[:180] or "—",
        "sender_name": _fix_mojibake(tx.sender_name or ""),
        "sender_iin_bin": tx.sender_iin_bin or "",
        "sender_account": tx.sender_account or "",
        "recipient_name": _fix_mojibake(tx.recipient_name or ""),
        "recipient_iin_bin": tx.recipient_iin_bin or "",
        "recipient_account": tx.recipient_account or "",
    }


async def _build_project_reconciliation_warnings(
    project_id: Optional[str],
    db: Optional[AsyncSession] = None,
) -> list[dict]:
    if not project_id:
        return []

    async def _load(session: AsyncSession) -> list[dict]:
        bank_rows = (
            await session.execute(
                select(Transaction)
                .where(
                    Transaction.project_id == project_id,
                    Transaction.source_bank != "esf",
                )
                .order_by(Transaction.operation_date.desc(), Transaction.date.desc())
            )
        ).scalars().all()
        esf_rows = (
            await session.execute(
                select(EsfRecord)
                .where(EsfRecord.project_id == project_id)
                .order_by(func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date).desc())
            )
        ).scalars().all()

        if not bank_rows or not esf_rows:
            return []

        warnings: list[dict] = []
        for esf in esf_rows:
            esf_refs = {
                _normalize_ref_token(esf.registration_number),
                _normalize_ref_token(esf.contract_number),
            }
            esf_refs = {ref for ref in esf_refs if ref}
            if not esf_refs:
                continue

            matched_bank_rows: list[Transaction] = []
            for tx in bank_rows:
                tx_refs = _extract_reference_tokens(tx.purpose, tx.raw_note, tx.operation_type)
                if not tx_refs:
                    continue
                if esf_refs & tx_refs:
                    matched_bank_rows.append(tx)

            if not matched_bank_rows:
                continue

            esf_amount = float(esf.total_amount or 0)
            bank_amount = sum(_tx_amount_value(tx) for tx in matched_bank_rows)
            difference = round(esf_amount - bank_amount, 2)
            if abs(difference) < 1:
                continue

            counterparty_name = _fix_mojibake(
                esf.buyer_name or esf.supplier_name or ""
            ) or "Не указан"
            counterparty_identifier = (
                _strip_iin_bin(esf.buyer_iin_bin)
                or _strip_iin_bin(esf.supplier_iin_bin)
                or "—"
            )
            references_label = ", ".join(
                ref for ref in [str(esf.contract_number or "").strip(), str(esf.registration_number or "").strip()] if ref
            )
            warnings.append(
                {
                    "code": "esf_bank_amount_mismatch",
                    "title": "Несоответствие суммы между ЭСФ и банковской оплатой",
                    "severity": "medium",
                    "summary": (
                        f"Для ЭСФ/договора {references_label or 'без номера'} в проекте найдена банковская оплата "
                        f"на {_format_money_label(bank_amount)}, тогда как сумма ЭСФ составляет {_format_money_label(esf_amount)}. "
                        f"Расхождение: {_format_money_label(abs(difference))}."
                    ),
                    "articles": ["сверка ЭСФ и банковских оплат"],
                    "indicators": [
                        {"label": "Сумма ЭСФ", "value": _format_money_label(esf_amount)},
                        {"label": "Сумма оплат в банке", "value": _format_money_label(bank_amount)},
                        {"label": "Расхождение", "value": _format_money_label(abs(difference))},
                        {"label": "Совпавших банковских операций", "value": str(len(matched_bank_rows))},
                    ],
                    "counterparties": [
                        {
                            "role": "Контрагент",
                            "name": counterparty_name,
                            "identifier": counterparty_identifier,
                            "transaction_count": len(matched_bank_rows),
                            "turnover": _format_money_label(bank_amount),
                            "articles": ["сверка ЭСФ и банковских оплат"],
                            "graph_iin_bin": counterparty_identifier if counterparty_identifier.isdigit() else None,
                        }
                    ],
                    "sample_transactions": [
                        _serialize_tx_as_warning_sample(tx)
                        for tx in sorted(matched_bank_rows, key=lambda item: _tx_amount_value(item), reverse=True)[:8]
                    ],
                }
            )

        return warnings

    if db is not None:
        return await _load(db)

    async with async_session() as session:
        return await _load(session)


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return 0.0

    s = s.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _safe_amount_tenge(amount_kzt: float, debit: float, credit: float) -> float:
    """
    Protect against parser misalignment where amount_kzt may accidentally contain
    huge header numbers (e.g., scientific-notation IIN from statement header).
    """
    base = max(float(debit or 0), float(credit or 0))
    amount = float(amount_kzt or 0)

    if amount <= 0:
        return base
    if base <= 0:
        return amount

    # If amount is far larger than transaction side amount, treat it as corrupted.
    ratio = amount / base if base > 0 else 1.0
    if ratio > 1000 or amount > 1e10:
        return base
    return amount


def _resolve_amount_tenge(currency: str, amount_currency: float, amount_kzt: float, debit: float, credit: float) -> float:
    normalized_currency = (currency or "KZT").strip().upper() or "KZT"
    raw_amount_kzt = float(amount_kzt or 0)
    if raw_amount_kzt > 0:
        return _safe_amount_tenge(raw_amount_kzt, debit, credit)

    base_amount = float(amount_currency or 0)
    if base_amount <= 0:
        base_amount = max(float(debit or 0), float(credit or 0))
    if base_amount <= 0:
        return 0.0

    rate = _KZT_CONVERSION_RATES.get(normalized_currency)
    if rate is None:
        return base_amount if normalized_currency == "KZT" else 0.0
    return round(base_amount * rate, 2)


def _to_str(value: object, max_len: int = 255) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s in {"'", '"'}:
        return ""
    return s[:max_len]


def _fix_mojibake(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        # Fix strings like "РїРµСЂРµРІРѕРґ" -> "перевод".
        fixed = s.encode("cp1251").decode("utf-8")
        if fixed:
            s = fixed
    except Exception:
        pass
    # Common parser-side artifact for "ё".
    s = s.replace("С‘", "ё").replace("Рµ", "е")
    return s


def _normalize_display_label(value: object) -> str:
    label = _fix_mojibake(value).strip()
    if not label:
        return ""
    if label.casefold() == "покупка в магазине":
        return "Оплата товаров и услуг"
    return label


def _derive_display_category(
    category: str,
    purpose_text: str,
    operation_type_raw: str,
    direction: str = "",
    *,
    transaction_category: str = "",
) -> str:
    base = _normalize_display_label(transaction_category or "")
    if base:
        return base
    base = _normalize_display_label(category or "")
    if base:
        return base
    return _normalize_display_label(
        _derive_category_from_core_row(base, purpose_text, operation_type_raw, direction)
    )


def _derived_category_expr():
    trimmed_category = func.nullif(func.trim(Transaction.transaction_category), "")
    trimmed_legacy_category = func.nullif(func.trim(Transaction.category), "")
    purpose = func.lower(func.coalesce(Transaction.purpose, ""))
    op_type = func.lower(func.coalesce(Transaction.operation_type, ""))
    direction = func.lower(func.coalesce(Transaction.direction, ""))
    return func.coalesce(
        trimmed_category,
        trimmed_legacy_category,
        case(
            (or_(purpose.like("%red.kz%"), purpose.like("%продаж%")), "Продажи Kaspi / Red"),
            (or_(purpose.like("%погаш%"), purpose.like("%кредит%")), "Погашение кредита"),
            (or_(purpose.like("%снятие%"), purpose.like("%cash%"), purpose.like("%atm%")), "Снятие наличных"),
            (or_(purpose.like("%ресайклер%"), purpose.like("%recycler%")), "Пополнение наличными"),
            (or_(purpose.like("%пополн%"), purpose.like("%взнос%"), purpose.like("%deposit%")), "Пополнение счёта"),
            (purpose.like("%рассроч%"), "Рассрочка Kaspi"),
            (purpose.like("%займ%"), "Выдача займа"),
            (and_(purpose.like("%перевод%"), purpose.like("%внутр%")), "Внутренние операции"),
            (or_(purpose.like("%перевод%"), op_type.like("%payment%")), "P2P перевод"),
            (purpose.like("%оплат%"), "Оплата услуг"),
            (direction == "credit", "Поступления"),
            (direction == "debit", "Расходы"),
            else_="Прочее",
        ),
    )


def _normalize_counterparty_identity(name: str, iin: str, account: str) -> tuple[str, str, str]:
    fixed_name = _fix_mojibake(name or "").strip()
    fixed_iin = re.sub(r"\D+", "", str(iin or ""))[:12]
    fixed_account = (str(account or "").strip().upper())[:20]

    unknown_tokens = {"", "UNKNOWN", "NULL", "NONE", "N/A", "-", "—", "000000000000"}
    if fixed_name.upper() in unknown_tokens:
        fixed_name = ""

    if not fixed_name:
        fixed_name = fixed_account or fixed_iin or ""

    return fixed_name, fixed_iin, fixed_account


def _resolve_counterparty_display_name(name: object, iin: object, account: object) -> str:
    normalized_name = _fix_mojibake(_to_str(name, 255)).strip()
    normalized_iin = re.sub(r"\D+", "", str(iin or ""))[:12]
    normalized_account = (str(account or "").strip().upper())[:64]
    return normalized_name or normalized_iin or normalized_account or ""


def _derive_category_from_core_row(sdp_name: str, purpose_text: str, operation_type_raw: str, direction: str) -> str:
    sdp = _fix_mojibake(sdp_name or "").strip()
    if sdp:
        return sdp

    purpose = _fix_mojibake(purpose_text or "").lower()
    op_type = _fix_mojibake(operation_type_raw or "").lower()
    direction_l = (direction or "").strip().lower()

    if "red.kz" in purpose or "продаж" in purpose:
        return "Продажи Kaspi / Red"
    if "погаш" in purpose or "кредит" in purpose:
        return "Погашение кредита"
    if "снятие" in purpose or "cash" in purpose or "atm" in purpose:
        return "Снятие наличных"
    if "\u0440\u0435\u0441\u0430\u0439\u043a\u043b\u0435\u0440" in purpose or "recycler" in purpose:
        return "Пополнение наличными"
    if "пополн" in purpose or "взнос" in purpose or "deposit" in purpose:
        return "Пополнение счёта"
    if "рассроч" in purpose:
        return "Рассрочка Kaspi"
    if "займ" in purpose:
        return "Выдача займа"
    if "перевод" in purpose and "внутр" in purpose:
        return "Внутренние операции"
    if "перевод" in purpose or "payment" in op_type:
        return "P2P перевод"
    if "оплат" in purpose:
        return "Оплата услуг"
    if direction_l == "credit":
        return "Поступления"
    if direction_l == "debit":
        return "Расходы"
    return "Прочее"


def _extract_transactions_from_transactions_core_csv(content: bytes) -> tuple[list[dict], int]:
    text = _decode_text_bytes(content)
    reader = csv.DictReader(StringIO(text))

    out: list[dict] = []
    skipped = 0

    for row in reader:
        if not row:
            continue

        op_dt = _parse_operation_datetime(row.get("operation_ts") or row.get("operation_date"))
        if op_dt is None:
            skipped += 1
            continue

        direction = (row.get("direction") or "").strip().lower()
        amount_currency = _to_float(row.get("amount_currency"))
        debit = _to_float(row.get("amount_debit"))
        credit = _to_float(row.get("amount_credit"))

        if debit <= 0 and credit <= 0 and amount_currency > 0:
            if direction == "debit":
                debit = amount_currency
            elif direction == "credit":
                credit = amount_currency

        currency = _to_str(row.get("currency"), 3).upper() or "KZT"
        amount_tenge = _resolve_amount_tenge(
            currency,
            amount_currency,
            _to_float(row.get("amount_kzt")),
            debit,
            credit,
        )
        if amount_tenge <= 0:
            skipped += 1
            continue

        sender_name, sender_iin, sender_account = _normalize_counterparty_identity(
            row.get("payer_name") or "",
            row.get("payer_iin_bin") or "",
            row.get("payer_account") or "",
        )
        recipient_name, recipient_iin, recipient_account = _normalize_counterparty_identity(
            row.get("receiver_name") or "",
            row.get("receiver_iin_bin") or "",
            row.get("receiver_account") or "",
        )

        purpose = _fix_mojibake(row.get("purpose_text") or row.get("raw_note") or "").strip()
        operation_type = _fix_mojibake(row.get("operation_type_raw") or "").strip()
        category = _derive_category_from_core_row(
            row.get("sdp_name") or "",
            purpose,
            operation_type,
            direction,
        )

        out.append({
            "date": op_dt,
            "sender_name": sender_name,
            "sender_iin_bin": sender_iin,
            "sender_account": sender_account,
            "recipient_name": recipient_name,
            "recipient_iin_bin": recipient_iin,
            "recipient_account": recipient_account,
            "purpose": purpose,
            "category": category,
            "operation_type": operation_type or direction,
            "currency": currency,
            "debit": debit,
            "credit": credit,
            "amount_tenge": amount_tenge,
        })

    return out, skipped


_ESF_DUPLICATE_HEADER_ALIASES = [
    "registration_number",
    "tax_authority_code",
    "esf_status",
    "issue_date",
    "turnover_date",
    "year",
    "supplier_iin_bin",
    "supplier_name",
    "supplier_address",
    "buyer_iin_bin",
    "buyer_name",
    "buyer_address",
    "country_code",
    "consignor_iin_bin",
    "consignor_name",
    "ship_from_address",
    "consignee_iin_bin",
    "consignee_name",
    "delivery_address",
    "contract_number",
    "contract_date",
    "payment_terms",
    "destination",
    "origin_sign",
    "tru_name",
    "tnved_code",
    "unit",
    "quantity",
    "price_without_vat",
    "price_with_vat",
    "cost_without_indirect_tax",
    "turnover_amount",
    "vat_rate",
    "vat_amount",
    "cost_with_indirect_tax",
    "total_amount",
    "currency_rate",
    "currency_code",
    "currency_name_ru",
]


def _rename_esf_headers(raw_headers: list[object]) -> list[str]:
    headers = [_fix_mojibake(_to_str(cell, 255)).strip() for cell in raw_headers]
    if len(headers) <= len(_ESF_DUPLICATE_HEADER_ALIASES):
        return _ESF_DUPLICATE_HEADER_ALIASES[: len(headers)]
    renamed = list(_ESF_DUPLICATE_HEADER_ALIASES)
    for idx in range(len(_ESF_DUPLICATE_HEADER_ALIASES), len(headers)):
        renamed.append(f"extra_{idx + 1}")
    return renamed


def _is_esf_header_row(row: list[object]) -> bool:
    normalized = [_fix_mojibake(_to_str(cell, 255)).strip().lower() for cell in row]
    return any("регистрационный номер эсф" in cell for cell in normalized) and any(
        "статус эсф" in cell for cell in normalized
    )


def _parse_esf_datetime(value: object) -> Optional[datetime]:
    parsed = _parse_operation_datetime(value)
    if parsed is not None:
        return parsed
    if value is None:
        return None
    text = _fix_mojibake(str(value)).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _strip_iin_bin(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    sci = raw.replace(",", ".")
    if re.search(r"[eE][+-]?\d+", sci):
        try:
            raw = str(int(round(float(sci))))
        except Exception:
            pass
    digits = re.sub(r"\D+", "", raw)
    return digits[:12]


def _pick_esf_amount(row: dict) -> float:
    for key in ("total_amount", "cost_with_indirect_tax", "turnover_amount", "cost_without_indirect_tax"):
        amount = _to_float(row.get(key))
        if amount > 0:
            return amount
    return 0.0


def _infer_esf_owner_side(rows: list[dict]) -> str:
    supplier_counts: dict[str, int] = {}
    buyer_counts: dict[str, int] = {}

    for row in rows:
        supplier_key = _strip_iin_bin(row.get("supplier_iin_bin")) or _normalize_text_key(row.get("supplier_name", ""))
        buyer_key = _strip_iin_bin(row.get("buyer_iin_bin")) or _normalize_text_key(row.get("buyer_name", ""))
        if supplier_key:
            supplier_counts[supplier_key] = supplier_counts.get(supplier_key, 0) + 1
        if buyer_key:
            buyer_counts[buyer_key] = buyer_counts.get(buyer_key, 0) + 1

    supplier_max = max(supplier_counts.values(), default=0)
    buyer_max = max(buyer_counts.values(), default=0)
    if buyer_max > supplier_max:
        return "buyer"
    return "supplier"


def _build_esf_transaction(row: dict, owner_side: str) -> Optional[dict]:
    op_dt = _parse_esf_datetime(row.get("turnover_date")) or _parse_esf_datetime(row.get("issue_date"))
    if op_dt is None:
        return None

    amount_tenge = _pick_esf_amount(row)
    if amount_tenge <= 0:
        return None

    supplier_name = _fix_mojibake(row.get("supplier_name") or "").strip()
    supplier_iin = _strip_iin_bin(row.get("supplier_iin_bin"))
    buyer_name = _fix_mojibake(row.get("buyer_name") or "").strip()
    buyer_iin = _strip_iin_bin(row.get("buyer_iin_bin"))
    purpose = _fix_mojibake(row.get("tru_name") or row.get("registration_number") or "").strip()
    esf_status = _fix_mojibake(row.get("esf_status") or "").strip()
    operation_type = f"ЭСФ {esf_status}".strip()
    currency = (_fix_mojibake(row.get("currency_code") or "").strip().upper() or "KZT")[:8]
    is_purchase = owner_side == "buyer"

    raw_bits = [
        _fix_mojibake(row.get("registration_number") or "").strip(),
        _fix_mojibake(row.get("contract_number") or "").strip(),
        _fix_mojibake(row.get("payment_terms") or "").strip(),
        _fix_mojibake(row.get("destination") or "").strip(),
        _fix_mojibake(row.get("tnved_code") or "").strip(),
    ]

    return {
        "date": op_dt,
        "sender_name": supplier_name,
        "sender_iin_bin": supplier_iin,
        "sender_account": "",
        "recipient_name": buyer_name,
        "recipient_iin_bin": buyer_iin,
        "recipient_account": "",
        "purpose": purpose,
        "category": "ЭСФ",
        "operation_type": operation_type or "ЭСФ",
        "currency": currency,
        "debit": amount_tenge if is_purchase else 0.0,
        "credit": amount_tenge if not is_purchase else 0.0,
        "amount_tenge": amount_tenge,
        "direction": "debit" if is_purchase else "credit",
        "raw_note": " | ".join(bit for bit in raw_bits if bit),
    }


def _extract_esf_rows_from_text(content: bytes) -> tuple[list[dict], int]:
    text = _decode_text_bytes(content)
    sample = text[:5000]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters="\t;,|").delimiter
    except Exception:
        delimiter = "\t" if "\t" in sample else ";"

    reader = csv.reader(StringIO(text, newline=""), delimiter=delimiter)
    rows = list(reader)
    header_idx = next((idx for idx, row in enumerate(rows) if _is_esf_header_row(list(row))), None)
    if header_idx is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не удалось найти заголовок ЭСФ в CSV")

    headers = _rename_esf_headers(list(rows[header_idx]))
    out: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if not any(str(cell or "").strip() for cell in row):
            continue
        out.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
    return out, 0


def _extract_esf_rows_from_workbook(content: bytes) -> tuple[list[dict], int]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = workbook[workbook.sheetnames[0]]
        scanned_rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        header_idx = next((idx for idx, row in enumerate(scanned_rows) if _is_esf_header_row(list(row or []))), None)
        if header_idx is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не удалось найти заголовок ЭСФ в Excel")
        headers = _rename_esf_headers(list(scanned_rows[header_idx]))
        out: list[dict] = []
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            if not any(str(cell or "").strip() for cell in row):
                continue
            out.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return out, 0
    finally:
        workbook.close()


def _extract_transactions_from_esf(content: bytes, filename: str) -> tuple[list[dict], int]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        rows, skipped = _extract_esf_rows_from_text(content)
    else:
        rows, skipped = _extract_esf_rows_from_workbook(content)

    owner_side = _infer_esf_owner_side(rows)
    out: list[dict] = []
    for row in rows:
        tx = _build_esf_transaction(row, owner_side)
        if tx is None:
            skipped += 1
            continue
        tx["raw_row_json"] = row
        out.append(tx)
    return out, skipped


def _extract_esf_records(content: bytes, filename: str) -> tuple[list[dict], str, int]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        rows, skipped = _extract_esf_rows_from_text(content)
    else:
        rows, skipped = _extract_esf_rows_from_workbook(content)
    return rows, _infer_esf_owner_side(rows), skipped


def _parse_optional_year(value: object) -> Optional[int]:
    try:
        numeric = int(float(str(value or "").strip()))
        return numeric if 1900 <= numeric <= 2100 else None
    except Exception:
        return None


def _parse_optional_decimal(value: object, precision: int = 2) -> Optional[float]:
    amount = _to_float(value)
    if amount == 0 and str(value or "").strip() == "":
        return None
    return round(amount, precision)


def _build_esf_record_payload(
    row: dict,
    *,
    owner_side: str,
    project_id: Optional[str],
    file_id: Optional[str],
    source_sheet: str,
    source_row_no: int,
) -> Optional[dict]:
    issue_dt = _parse_esf_datetime(row.get("issue_date"))
    turnover_dt = _parse_esf_datetime(row.get("turnover_date"))
    total_amount = _pick_esf_amount(row)
    if turnover_dt is None and issue_dt is None:
        return None
    if total_amount <= 0:
        return None

    normalized_row = {
        key: _fix_mojibake(value).strip() if isinstance(value, str) else value
        for key, value in row.items()
    }

    signature_source = json.dumps(
        {
            "project_id": project_id,
            "registration_number": normalized_row.get("registration_number"),
            "issue_date": issue_dt.isoformat() if issue_dt else None,
            "turnover_date": turnover_dt.isoformat() if turnover_dt else None,
            "supplier_iin_bin": _strip_iin_bin(normalized_row.get("supplier_iin_bin")),
            "buyer_iin_bin": _strip_iin_bin(normalized_row.get("buyer_iin_bin")),
            "tru_name": normalized_row.get("tru_name"),
            "total_amount": round(total_amount, 2),
        },
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )

    return {
        "id": str(uuid4()),
        "project_id": project_id,
        "file_id": file_id,
        "source_sheet": source_sheet,
        "source_row_no": source_row_no,
        "row_hash": hashlib.sha256(signature_source.encode("utf-8")).hexdigest(),
        "esf_direction": "purchase" if owner_side == "buyer" else "sale",
        "registration_number": _fix_mojibake(normalized_row.get("registration_number") or "").strip(),
        "tax_authority_code": _to_str(normalized_row.get("tax_authority_code"), 64),
        "esf_status": _to_str(normalized_row.get("esf_status"), 64),
        "issue_date": issue_dt,
        "turnover_date": turnover_dt,
        "year": _parse_optional_year(normalized_row.get("year")),
        "supplier_iin_bin": _strip_iin_bin(normalized_row.get("supplier_iin_bin")),
        "supplier_name": _to_str(normalized_row.get("supplier_name"), 500),
        "supplier_address": _to_str(normalized_row.get("supplier_address"), 1000),
        "buyer_iin_bin": _strip_iin_bin(normalized_row.get("buyer_iin_bin")),
        "buyer_name": _to_str(normalized_row.get("buyer_name"), 500),
        "buyer_address": _to_str(normalized_row.get("buyer_address"), 1000),
        "country_code": _to_str(normalized_row.get("country_code"), 16),
        "consignor_iin_bin": _strip_iin_bin(normalized_row.get("consignor_iin_bin")),
        "consignor_name": _to_str(normalized_row.get("consignor_name"), 500),
        "ship_from_address": _to_str(normalized_row.get("ship_from_address"), 1000),
        "consignee_iin_bin": _strip_iin_bin(normalized_row.get("consignee_iin_bin")),
        "consignee_name": _to_str(normalized_row.get("consignee_name"), 500),
        "delivery_address": _to_str(normalized_row.get("delivery_address"), 1000),
        "contract_number": _to_str(normalized_row.get("contract_number"), 255),
        "contract_date": _parse_esf_datetime(normalized_row.get("contract_date")),
        "payment_terms": _to_str(normalized_row.get("payment_terms"), 255),
        "destination": _to_str(normalized_row.get("destination"), 255),
        "origin_sign": _to_str(normalized_row.get("origin_sign"), 255),
        "tru_name": _to_str(normalized_row.get("tru_name"), 1000),
        "tnved_code": _to_str(normalized_row.get("tnved_code"), 255),
        "unit": _to_str(normalized_row.get("unit"), 128),
        "quantity": _parse_optional_decimal(normalized_row.get("quantity"), 4),
        "price_without_vat": _parse_optional_decimal(normalized_row.get("price_without_vat"), 2),
        "price_with_vat": _parse_optional_decimal(normalized_row.get("price_with_vat"), 2),
        "cost_without_indirect_tax": _parse_optional_decimal(normalized_row.get("cost_without_indirect_tax"), 2),
        "turnover_amount": _parse_optional_decimal(normalized_row.get("turnover_amount"), 2),
        "vat_rate": _parse_optional_decimal(normalized_row.get("vat_rate"), 4),
        "vat_amount": _parse_optional_decimal(normalized_row.get("vat_amount"), 2),
        "cost_with_indirect_tax": _parse_optional_decimal(normalized_row.get("cost_with_indirect_tax"), 2),
        "total_amount": _parse_optional_decimal(normalized_row.get("total_amount"), 2) or round(total_amount, 2),
        "currency_rate": _parse_optional_decimal(normalized_row.get("currency_rate"), 6),
        "currency_code": _to_str(normalized_row.get("currency_code"), 16).upper() or "KZT",
        "currency_name_ru": _to_str(normalized_row.get("currency_name_ru"), 255),
        "raw_row_json": normalized_row,
    }


def _build_esf_shadow_transaction_payload(
    esf_payload: dict,
    *,
    uploader_email: str,
) -> dict:
    tx_id = str(esf_payload["id"])
    amount = float(esf_payload.get("total_amount") or 0)
    is_purchase = (esf_payload.get("esf_direction") or "").strip().lower() == "purchase"
    direction = "debit" if is_purchase else "credit"
    operation_type = f"ЭСФ {(esf_payload.get('esf_status') or '').strip()}".strip()
    category = "Приобретение" if is_purchase else "Реализация"
    purpose_parts = [
        esf_payload.get("tru_name") or "",
        esf_payload.get("registration_number") or "",
        esf_payload.get("contract_number") or "",
    ]
    purpose_text = " | ".join(part.strip() for part in purpose_parts if str(part or "").strip())
    semantic_parts = [
        "esf",
        operation_type,
        category,
        esf_payload.get("registration_number") or "",
        esf_payload.get("tru_name") or "",
        esf_payload.get("supplier_name") or "",
        esf_payload.get("buyer_name") or "",
        esf_payload.get("contract_number") or "",
    ]
    semantic_text = " | ".join(part.strip() for part in semantic_parts if str(part or "").strip())

    return {
        "tx_id": tx_id,
        "file_id": esf_payload.get("file_id"),
        "statement_id": None,
        "format_id": None,
        "project_id": esf_payload.get("project_id"),
        "source_bank": "esf",
        "source_sheet": esf_payload.get("source_sheet"),
        "source_block_id": None,
        "source_row_no": esf_payload.get("source_row_no"),
        "row_hash": f"esf-shadow:{esf_payload.get('row_hash')}",
        "date": esf_payload.get("turnover_date") or esf_payload.get("issue_date"),
        "operation_date": (esf_payload.get("turnover_date") or esf_payload.get("issue_date")).date() if (esf_payload.get("turnover_date") or esf_payload.get("issue_date")) else None,
        "currency": (esf_payload.get("currency_code") or "KZT").strip().upper(),
        "amount_currency": amount,
        "amount_tenge": amount,
        "credit": 0 if is_purchase else amount,
        "debit": amount if is_purchase else 0,
        "direction": direction,
        "operation_type": operation_type,
        "category": "ЭСФ",
        "purpose_code": None,
        "purpose": purpose_text,
        "raw_note": esf_payload.get("contract_number"),
        "sender_name": esf_payload.get("supplier_name"),
        "sender_iin_bin": esf_payload.get("supplier_iin_bin"),
        "payer_residency": esf_payload.get("supplier_address"),
        "payer_bank": None,
        "sender_account": "",
        "recipient_name": esf_payload.get("buyer_name"),
        "recipient_iin_bin": esf_payload.get("buyer_iin_bin"),
        "receiver_residency": esf_payload.get("buyer_address"),
        "receiver_bank": None,
        "recipient_account": "",
        "confidence_score": 1.0,
        "parse_warnings": None,
        "raw_row_json": esf_payload.get("raw_row_json"),
        "transaction_category": category,
        "category_confidence": None,
        "category_source": "esf",
        "category_rule_id": None,
        "needs_review": False,
        "semantic_text": semantic_text,
    }


def _to_mojibake(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    try:
        return s.encode("utf-8").decode("cp1251")
    except Exception:
        return s


def _norm_header(value: object) -> str:
    s = _fix_mojibake(str(value or "")).strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def _pick_index(index_map: dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        candidates = (
            _norm_header(alias),
            _norm_header(_fix_mojibake(alias)),
            _norm_header(_to_mojibake(alias)),
        )
        for candidate in candidates:
            if candidate in index_map:
                return index_map[candidate]
    return None


def _get_cell(row: tuple, idx: Optional[int]) -> object:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _extract_sheet_client_fallback(rows: list[tuple]) -> tuple[str, str, str]:
    """
    Try to read owner identity from statement header blocks:
    - Клиент:
    - ИИН/БИН:
    - Счет:
    """
    client_name = ""
    client_iin = ""
    client_account = ""

    def _clean_iin(value: str) -> str:
        s = str(value or "").strip().replace("\u00a0", "").replace(" ", "")
        if not s:
            return ""
        sci = s.replace(",", ".")
        if re.search(r"[eE][+-]?\d+", sci):
            try:
                return str(int(round(float(sci))))
            except Exception:
                pass
        return "".join(ch for ch in s if ch.isdigit())

    scan_rows = rows[:40]
    for row in scan_rows:
        if not row:
            continue
        label = _fix_mojibake(_to_str(row[0], 128)).strip().lower().replace("ё", "е")
        if not label:
            continue

        # Usually value is in column B/C in raw exports.
        value = ""
        for idx in (1, 2, 3):
            if idx < len(row):
                cell_val = _fix_mojibake(_to_str(row[idx], 255)).strip()
                if cell_val:
                    value = cell_val
                    break
        if not value:
            continue

        if "клиент" in label and not client_name:
            client_name = value.strip('"').strip("'")
        elif ("иин/бин" in label or "иин" in label or "бин" in label) and not client_iin:
            client_iin = _clean_iin(value)[:12]
        elif ("счет" in label or "счет:" in label or "счета" in label) and not client_account:
            client_account = value

    return client_name, client_iin, client_account


def _is_incoming(direction_value: str, credit: float) -> bool:
    d = _fix_mojibake((direction_value or "").strip()).lower()
    if "вход" in d or "incoming" in d:
        return True
    if "исход" in d or "outgoing" in d:
        return False
    return credit > 0


def _get_parser_url(parser_type: str) -> str:
    parser_type = (parser_type or PARSER_KASPI).strip().lower()
    if parser_type == PARSER_HALYK:
        base = (settings.HALYK_PARSER_API_URL or "").strip()
        if not base:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="HALYK_PARSER_API_URL is not configured",
            )
        return base.rstrip("/") + "/parse"

    if parser_type == PARSER_KASPI:
        base = (
            (settings.KASPI_PARSER_API_URL or "").strip()
            or (settings.BANK_PARSER_API_URL or "").strip()
            or settings.PARSER_API_URL
        )
        return base.rstrip("/") + "/parse"

    if parser_type == PARSER_KASPI_LEGACY:
        base = (
            (settings.BANK_PARSER_API_URL or "").strip()
            or settings.PARSER_API_URL
        )
        return base.rstrip("/")

    if parser_type != PARSER_BANK:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unknown parser_type. Use smart_parser, kaspi, kaspi_parser, bank_parser, halyk_parser or transactions_core",
        )

    base = (settings.BANK_PARSER_API_URL or "").strip() or settings.PARSER_API_URL
    return base.rstrip("/") + "/parse"


def _identity_quality(name: str, iin: str, account: str) -> int:
    score = 0
    n = (name or "").strip().upper()
    i = (iin or "").strip()
    a = (account or "").strip()
    if n and n != "UNKNOWN":
        score += 2
    if i and i not in {"0", "000000000000"}:
        score += 2
    if a:
        score += 1
    return score


def _normalize_id_value(value: str) -> str:
    return "".join(ch for ch in (value or "").strip().upper() if ch.isalnum())


def _normalize_text_key(value: str) -> str:
    return re.sub(r"\s+", " ", _fix_mojibake(value or "").strip().lower())


def _tx_quality(tx: dict) -> int:
    score = 0
    score += _identity_quality(tx.get("sender_name", ""), tx.get("sender_iin_bin", ""), tx.get("sender_account", ""))
    score += _identity_quality(tx.get("recipient_name", ""), tx.get("recipient_iin_bin", ""), tx.get("recipient_account", ""))
    if (tx.get("purpose") or "").strip():
        score += 1
    if (tx.get("category") or "").strip():
        score += 1
    if (tx.get("operation_type") or "").strip():
        score += 1
    return score


def _tx_signature(tx: dict) -> tuple:
    dt: datetime = tx["date"]
    return (
        dt.isoformat(timespec="seconds"),
        (tx.get("currency") or "KZT").strip().upper(),
        round(float(tx.get("debit") or 0), 2),
        round(float(tx.get("credit") or 0), 2),
        round(float(tx.get("amount_tenge") or 0), 2),
        _normalize_id_value(tx.get("sender_iin_bin", "")),
        _normalize_id_value(tx.get("sender_account", "")),
        _normalize_text_key(tx.get("sender_name", ""))[:120],
        _normalize_id_value(tx.get("recipient_iin_bin", "")),
        _normalize_id_value(tx.get("recipient_account", "")),
        _normalize_text_key(tx.get("recipient_name", ""))[:120],
        _normalize_text_key(tx.get("purpose", ""))[:400],
        _normalize_text_key(tx.get("category", ""))[:120],
        _normalize_text_key(tx.get("operation_type", ""))[:120],
    )


def _build_core_transaction_payload(
    tx: dict,
    *,
    uploader_email: str,
    project_id: Optional[str],
    source_bank: str,
    source_sheet: str,
    source_row_no: int,
    file_id: Optional[str] = None,
    statement_id: Optional[str] = None,
    format_id: Optional[str] = None,
    source_block_id: int = 1,
    raw_row_json: Optional[dict] = None,
) -> dict:
    def _jsonable(value):
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, dict):
            return {str(k): _jsonable(v) for k, v in value.items()}
        if isinstance(value, (list, tuple)):
            return [_jsonable(v) for v in value]
        return value

    op_dt: datetime = tx["date"]
    # Ensure amount_tenge (amount_kzt) is always a valid float and not None/NaN
    def _safe_float(val):
        try:
            f = float(val)
            if f != f:  # check for NaN
                return 0.0
            return f
        except Exception:
            return 0.0

    # Robustly get amount_tenge (amount_kzt) from tx dict
    def _get_amount_tenge(tx):
        for key in ("amount_tenge", "amount_kzt"):
            v = tx.get(key)
            if v is not None:
                return _safe_float(v)
        return 0.0

    payload = {
        "id": str(tx.get("id") or uuid4()),
        "file_id": file_id,
        "statement_id": statement_id,
        "format_id": format_id,
        "project_id": project_id,
        "source_bank": source_bank,
        "source_sheet": (source_sheet or "")[:255] or None,
        "source_block_id": source_block_id,
        "source_row_no": source_row_no,
        "date": op_dt,
        "operation_date": op_dt.date(),
        "currency": (tx.get("currency") or "KZT").strip().upper(),
        "amount_currency": float(max(float(tx.get("debit") or 0), float(tx.get("credit") or 0))),
        "amount_tenge": _get_amount_tenge(tx),
        "credit": float(tx.get("credit") or 0),
        "debit": float(tx.get("debit") or 0),
        "direction": "credit" if float(tx.get("credit") or 0) > 0 else "debit",
        "operation_type": _fix_mojibake(tx.get("operation_type") or ""),
        "category": _fix_mojibake(tx.get("category") or ""),
        "purpose_code": None,
        "purpose": _fix_mojibake(tx.get("purpose") or ""),
        "raw_note": _fix_mojibake(tx.get("purpose") or ""),
        "sender_name": _resolve_counterparty_display_name(
            tx.get("sender_name"),
            tx.get("sender_iin_bin"),
            tx.get("sender_account"),
        ),
        "sender_iin_bin": (tx.get("sender_iin_bin") or "")[:32],
        "payer_residency": None,
        "payer_bank": None,
        "sender_account": (tx.get("sender_account") or "")[:64],
        "recipient_name": _resolve_counterparty_display_name(
            tx.get("recipient_name"),
            tx.get("recipient_iin_bin"),
            tx.get("recipient_account"),
        ),
        "recipient_iin_bin": (tx.get("recipient_iin_bin") or "")[:32],
        "receiver_residency": None,
        "receiver_bank": None,
        "recipient_account": (tx.get("recipient_account") or "")[:64],
        "confidence_score": 1.0,
        "parse_warnings": None,
        "raw_row_json": _jsonable(raw_row_json) if raw_row_json is not None else None,
    }

    sig_source = json.dumps(
        {
            "date": op_dt.isoformat(timespec="seconds"),
            "currency": payload["currency"],
            "debit": payload["debit"],
            "credit": payload["credit"],
            "sender_iin_bin": payload["sender_iin_bin"],
            "sender_account": payload["sender_account"],
            "recipient_iin_bin": payload["recipient_iin_bin"],
            "recipient_account": payload["recipient_account"],
            "purpose": payload["purpose"],
            "category": payload["category"],
            "operation_type": payload["operation_type"],
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    payload["row_hash"] = hashlib.sha256(sig_source.encode("utf-8")).hexdigest()
    return payload


def _make_xlsx_with_single_sheet(sheet_title: str, rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    for row in rows:
        ws.append(list(row))
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _extract_transactions_from_workbook(workbook, parser_type: str) -> tuple[list[dict], int]:
    parser_type = (parser_type or "").strip().lower()
    out: list[dict] = []
    skipped = 0

    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        sheet_client_name, sheet_client_iin, sheet_client_account = _extract_sheet_client_fallback(rows)

        header_norm = [_norm_header(c) for c in rows[0]]
        index = {name: i for i, name in enumerate(header_norm)}

        if parser_type in {PARSER_HALYK, PARSER_KASPI}:
            idx_date = _pick_index(index, "РґР°С‚Р° РѕРїРµСЂР°С†РёРё", "operation_date")
            idx_currency = _pick_index(index, "РІР°Р»СЋС‚Р° РѕРїРµСЂР°С†РёРё", "РІР°Р»СЋС‚Р°", "currency")
            idx_debit = _pick_index(index, "СЃСѓРјРјР° РґРµР±РµС‚", "РґРµР±РµС‚", "debit")
            idx_credit = _pick_index(index, "СЃСѓРјРјР° РєСЂРµРґРёС‚", "РєСЂРµРґРёС‚", "credit")
            idx_amount = _pick_index(index, "СЃСѓРјРјР° РІ С‚РµРЅРіРµ", "amount_kzt")
            idx_direction = _pick_index(index, "РЅР°РїСЂР°РІР»РµРЅРёРµ", "direction")
            idx_owner_name = _pick_index(index, "РІР»Р°РґРµР»РµС† СЃС‡РµС‚Р°", "owner_name")
            idx_owner_iin = _pick_index(index, "РёРёРЅ/Р±РёРЅ РІР»Р°РґРµР»СЊС†Р°", "owner_iin")
            idx_owner_acc = _pick_index(index, "РЅРѕРјРµСЂ СЃС‡РµС‚Р° РІР»Р°РґРµР»СЊС†Р°", "РЅРѕРјРµСЂ СЃС‡РµС‚Р° РІР»Р°РґРµР»СЊС†Р°", "РЅРѕРјРµСЂ СЃС‡С‘С‚Р° РІР»Р°РґРµР»СЊС†Р°", "owner_account")
            idx_cp_name = _pick_index(index, "РєРѕРЅС‚СЂР°РіРµРЅС‚", "recipient", "counterparty")
            idx_cp_iin = _pick_index(index, "РёРёРЅ/Р±РёРЅ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "iin_recipient", "counterparty_iin")
            idx_cp_acc = _pick_index(index, "СЃС‡РµС‚ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "СЃС‡С‘С‚ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "account_recipient", "counterparty_account")
            idx_purpose = _pick_index(index, "РЅР°Р·РЅР°С‡РµРЅРёРµ РїР»Р°С‚РµР¶Р°", "transfer_purpose", "purpose")
            idx_category = _pick_index(index, "РєР°С‚РµРіРѕСЂРёСЏ", "category")
            idx_operation_type = _pick_index(
                index,
                "operation_type",
                "document_category",
                "op_type",
                "type",
                "РўРёРї РѕРїРµСЂР°С†РёРё",
                "РІРёРґС‹ РѕРїРµСЂР°С†РёРё (РєР°С‚РµРіРѕСЂРёСЏ РґРѕРєСѓРјРµРЅС‚Р°)",
            )
            idx_kaspi_payer_name = _pick_index(index, "payer_name", "payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_payer_iin = _pick_index(index, "iin_payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_payer_acc = _pick_index(index, "account_payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_name = _pick_index(index, "recipient_name", "recipient") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_iin = _pick_index(index, "iin_recipient") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_acc = _pick_index(index, "account_recipient") if parser_type == PARSER_KASPI else None

            if idx_date is None or idx_currency is None or idx_debit is None or idx_credit is None:
                continue

            for row in rows[1:]:
                op_dt = _parse_operation_datetime(_get_cell(row, idx_date))
                if op_dt is None:
                    skipped += 1
                    continue

                debit = _to_float(_get_cell(row, idx_debit))
                credit = _to_float(_get_cell(row, idx_credit))
                amount_kzt = _to_float(_get_cell(row, idx_amount))
                amount_tenge = _safe_amount_tenge(amount_kzt, debit, credit)
                if amount_tenge <= 0:
                    skipped += 1
                    continue

                owner_name = _to_str(_get_cell(row, idx_owner_name), 255)
                owner_iin = _to_str(_get_cell(row, idx_owner_iin), 12)
                owner_acc = _to_str(_get_cell(row, idx_owner_acc), 20)
                if not owner_name or owner_name.upper() == "UNKNOWN":
                    owner_name = sheet_client_name or owner_name
                if (not owner_iin or owner_iin == "000000000000") and sheet_client_iin:
                    owner_iin = sheet_client_iin
                if not owner_acc and sheet_client_account:
                    owner_acc = sheet_client_account

                cp_name = _to_str(_get_cell(row, idx_cp_name), 255)
                cp_iin = _to_str(_get_cell(row, idx_cp_iin), 12)
                cp_acc = _to_str(_get_cell(row, idx_cp_acc), 20)
                direction_value = _to_str(_get_cell(row, idx_direction), 64)
                incoming = _is_incoming(direction_value, credit)

                kaspi_payer_name = _to_str(_get_cell(row, idx_kaspi_payer_name), 255) if idx_kaspi_payer_name is not None else ""
                kaspi_payer_iin = _to_str(_get_cell(row, idx_kaspi_payer_iin), 12) if idx_kaspi_payer_iin is not None else ""
                kaspi_payer_acc = _to_str(_get_cell(row, idx_kaspi_payer_acc), 20) if idx_kaspi_payer_acc is not None else ""
                kaspi_recipient_name = _to_str(_get_cell(row, idx_kaspi_recipient_name), 255) if idx_kaspi_recipient_name is not None else ""
                kaspi_recipient_iin = _to_str(_get_cell(row, idx_kaspi_recipient_iin), 12) if idx_kaspi_recipient_iin is not None else ""
                kaspi_recipient_acc = _to_str(_get_cell(row, idx_kaspi_recipient_acc), 20) if idx_kaspi_recipient_acc is not None else ""
                has_kaspi_explicit_parties = any(
                    [kaspi_payer_name, kaspi_payer_iin, kaspi_payer_acc, kaspi_recipient_name, kaspi_recipient_iin, kaspi_recipient_acc]
                )

                if parser_type == PARSER_KASPI and has_kaspi_explicit_parties:
                    sender_name = kaspi_payer_name
                    sender_iin = kaspi_payer_iin
                    sender_acc = kaspi_payer_acc
                    recipient_name = kaspi_recipient_name
                    recipient_iin = kaspi_recipient_iin
                    recipient_acc = kaspi_recipient_acc
                else:
                    sender_name = cp_name if incoming else owner_name
                    sender_iin = cp_iin if incoming else owner_iin
                    sender_acc = cp_acc if incoming else owner_acc
                    recipient_name = owner_name if incoming else cp_name
                    recipient_iin = owner_iin if incoming else cp_iin
                    recipient_acc = owner_acc if incoming else cp_acc

                if (not sender_name or sender_name.upper() == "UNKNOWN") and sheet_client_name:
                    sender_name = sheet_client_name
                if (not sender_iin or sender_iin == "000000000000") and sheet_client_iin:
                    sender_iin = sheet_client_iin
                if not sender_acc and sheet_client_account:
                    sender_acc = sheet_client_account

                out.append({
                    "date": op_dt,
                    "sender_name": sender_name or "UNKNOWN",
                    "sender_iin_bin": sender_iin or "000000000000",
                    "sender_account": sender_acc or "",
                    "recipient_name": recipient_name or "UNKNOWN",
                    "recipient_iin_bin": recipient_iin or "000000000000",
                    "recipient_account": recipient_acc or "",
                    "purpose": _to_str(_get_cell(row, idx_purpose), 1000),
                    "category": _fix_mojibake(_to_str(_get_cell(row, idx_category), 255)),
                    "operation_type": _fix_mojibake(_to_str(_get_cell(row, idx_operation_type), 255)) or _fix_mojibake(direction_value),
                    "currency": _to_str(_get_cell(row, idx_currency), 3).upper() or "KZT",
                    "debit": debit,
                    "credit": credit,
                    "amount_tenge": amount_tenge,
                })
        else:
            idx_date = _pick_index(index, "operation_date")
            idx_currency = _pick_index(index, "currency")
            idx_debit = _pick_index(index, "debit")
            idx_credit = _pick_index(index, "credit")
            if idx_date is None or idx_currency is None or idx_debit is None or idx_credit is None:
                continue

            idx_amount = _pick_index(index, "amount_kzt")
            idx_payer = _pick_index(index, "payer")
            idx_iin_payer = _pick_index(index, "iin_payer")
            idx_acc_payer = _pick_index(index, "account_payer")
            idx_recipient = _pick_index(index, "recipient")
            idx_iin_recipient = _pick_index(index, "iin_recipient")
            idx_acc_recipient = _pick_index(index, "account_recipient")
            idx_purpose = _pick_index(index, "transfer_purpose")
            idx_category = _pick_index(index, "category")
            idx_operation_type = _pick_index(
                index,
                "operation_type",
                "document_category",
                "op_type",
                "type",
                "РўРёРї РѕРїРµСЂР°С†РёРё",
            )

            for row in rows[1:]:
                op_dt = _parse_operation_datetime(_get_cell(row, idx_date))
                if op_dt is None:
                    skipped += 1
                    continue

                debit = _to_float(_get_cell(row, idx_debit))
                credit = _to_float(_get_cell(row, idx_credit))
                amount_kzt = _to_float(_get_cell(row, idx_amount))
                amount_tenge = _safe_amount_tenge(amount_kzt, debit, credit)
                if amount_tenge <= 0:
                    skipped += 1
                    continue

                out.append({
                    "date": op_dt,
                    "sender_name": _to_str(_get_cell(row, idx_payer), 255),
                    "sender_iin_bin": _to_str(_get_cell(row, idx_iin_payer), 12),
                    "sender_account": _to_str(_get_cell(row, idx_acc_payer), 20),
                    "recipient_name": _to_str(_get_cell(row, idx_recipient), 255),
                    "recipient_iin_bin": _to_str(_get_cell(row, idx_iin_recipient), 12),
                    "recipient_account": _to_str(_get_cell(row, idx_acc_recipient), 20),
                    "purpose": _to_str(_get_cell(row, idx_purpose), 1000),
                    "category": _fix_mojibake(_to_str(_get_cell(row, idx_category), 255)),
                    "operation_type": _fix_mojibake(_to_str(_get_cell(row, idx_operation_type), 255)),
                    "currency": _to_str(_get_cell(row, idx_currency), 3).upper() or "KZT",
                    "debit": debit,
                    "credit": credit,
                    "amount_tenge": amount_tenge,
                })

    return out, skipped


def _require_admin(authorization: Optional[str]) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing bearer token")

    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = decode_access_token(token)
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    if payload.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    return payload


def _parser_type_to_source_bank(parser_type: str) -> Optional[str]:
    normalized = (parser_type or "").strip().lower()
    if normalized in {PARSER_KASPI, PARSER_KASPI_LEGACY}:
        return "kaspi"
    if normalized == PARSER_HALYK:
        return "halyk"
    if normalized == PARSER_ESF:
        return "esf"
    if normalized in {PARSER_SMART, PARSER_BANK}:
        return None
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unknown parser_type. Use smart_parser, kaspi, kaspi_parser, bank_parser, halyk_parser, esf or transactions_core",
    )


def _run_ingestion_pipeline(file_path: str, source_bank: Optional[str], project_id: Optional[str]):
    with IngestionPipeline() as pipeline:
        return pipeline.ingest_file(file_path, source_bank=source_bank, project_id=project_id)


async def _finalize_ingestion_import(file_id: Optional[str], uploader_email: str, project_id: Optional[str]) -> int:
    if not file_id:
        return 0

    async with async_session() as session:
        uploaded_tx_ids = list(
            (
                await session.execute(
                    select(Transaction.id).where(
                        Transaction.file_id == file_id,
                        Transaction.project_id == project_id,
                    )
                )
            ).scalars()
        )
        if not uploaded_tx_ids:
            return 0

        existing_meta = set(
            (
                await session.execute(
                    select(TransactionUploadMeta.tx_id).where(TransactionUploadMeta.tx_id.in_(uploaded_tx_ids))
                )
            ).scalars()
        )

        for tx_id in uploaded_tx_ids:
            if tx_id in existing_meta:
                continue
            session.add(
                TransactionUploadMeta(
                    tx_id=tx_id,
                    project_id=project_id,
                    uploaded_by_email=uploader_email or "",
                    created_at=datetime.utcnow(),
                )
            )

        if len(existing_meta) != len(uploaded_tx_ids):
            await session.commit()

        return len(uploaded_tx_ids)


def _text_match_conditions(column, query: str):
    q = (query or "").strip()
    if not q:
        return None
    variants = {q, q.lower(), q.upper(), q.capitalize(), q.title()}
    like_conds = [column.ilike(f"%{v}%") for v in variants if v]
    return or_(*like_conds) if like_conds else None


def _non_empty_text_condition(column):
    return func.nullif(func.trim(func.coalesce(column, "")), "").isnot(None)


def _meaningful_transaction_condition():
    return or_(
        Transaction.date.isnot(None),
        Transaction.operation_date.isnot(None),
        func.coalesce(Transaction.amount_tenge, 0) > 0,
        func.coalesce(Transaction.debit, 0) > 0,
        func.coalesce(Transaction.credit, 0) > 0,
        _non_empty_text_condition(Transaction.sender_name),
        _non_empty_text_condition(Transaction.sender_account),
        _non_empty_text_condition(Transaction.recipient_name),
        _non_empty_text_condition(Transaction.recipient_account),
        _non_empty_text_condition(Transaction.purpose),
        _non_empty_text_condition(Transaction.currency),
    )


def _build_transaction_where_clause(
    project_id: str,
    date: Optional[str],
    category: Optional[str],
    search: Optional[str],
    min_amount: Optional[float],
    max_amount: Optional[float],
    currency: Optional[str],
    sender: Optional[str],
    recipient: Optional[str],
    include_esf: bool = False,
):
    conditions = [
        Transaction.project_id == project_id,
        _meaningful_transaction_condition(),
    ]
    if not include_esf:
        conditions.append(Transaction.source_bank != "esf")
    if date:
        dt = _parse_date(date)
        conditions.append(
            or_(
                cast(func.timezone("UTC", Transaction.date), Date) == dt.date(),
                Transaction.operation_date == dt.date(),
            )
        )

    if category:
        category_q = category.strip()
        category_mojibake = _to_mojibake(category_q)
        category_expr = _derived_category_expr()
        cond_main = _text_match_conditions(category_expr, category_q)
        if cond_main is not None:
            if category_mojibake and category_mojibake != category_q:
                cond_moji = _text_match_conditions(category_expr, category_mojibake)
                if cond_moji is not None:
                    conditions.append(or_(cond_main, cond_moji))
                else:
                    conditions.append(cond_main)
            else:
                conditions.append(cond_main)

    if search:
        search_cond = _text_match_conditions(Transaction.purpose, search)
        if search_cond is not None:
            conditions.append(search_cond)
    if min_amount is not None:
        conditions.append(Transaction.amount_tenge >= min_amount)
    if max_amount is not None:
        conditions.append(Transaction.amount_tenge <= max_amount)
    if currency:
        conditions.append(Transaction.currency == currency.upper())
    if sender:
        sender_q = sender.strip()
        sender_name_cond = _text_match_conditions(Transaction.sender_name, sender_q)
        sender_match_list = [
            Transaction.sender_iin_bin.ilike(f"%{sender_q}%"),
            Transaction.sender_account.ilike(f"%{sender_q}%"),
        ]
        if sender_name_cond is not None:
            sender_match_list.insert(0, sender_name_cond)
        conditions.append(
            or_(*sender_match_list)
        )
    if recipient:
        recipient_q = recipient.strip()
        recipient_name_cond = _text_match_conditions(Transaction.recipient_name, recipient_q)
        recipient_match_list = [
            Transaction.recipient_iin_bin.ilike(f"%{recipient_q}%"),
            Transaction.recipient_account.ilike(f"%{recipient_q}%"),
        ]
        if recipient_name_cond is not None:
            recipient_match_list.insert(0, recipient_name_cond)
        conditions.append(
            or_(*recipient_match_list)
        )
    return and_(*conditions) if conditions else True


def _build_transaction_order_by(sort_by: str, sort_dir: str):
    key = (sort_by or "date").strip()
    direction = (sort_dir or "desc").strip().lower()
    is_desc = direction != "asc"

    text_exprs = {
        "category": func.lower(func.coalesce(_derived_category_expr(), "")),
        "operationType": func.lower(func.coalesce(Transaction.operation_type, "")),
        "purpose": func.lower(func.coalesce(Transaction.purpose, "")),
        "uploadedBy": func.lower(func.coalesce(TransactionUploadMeta.uploaded_by_email, "")),
        "sender": func.lower(
            func.coalesce(
                func.nullif(func.trim(Transaction.sender_name), ""),
                func.nullif(func.trim(Transaction.sender_account), ""),
                Transaction.sender_iin_bin,
                "",
            )
        ),
        "recipient": func.lower(
            func.coalesce(
                func.nullif(func.trim(Transaction.recipient_name), ""),
                func.nullif(func.trim(Transaction.recipient_account), ""),
                Transaction.recipient_iin_bin,
                "",
            )
        ),
    }

    numeric_exprs = {
        "date": Transaction.date,
        "debit": Transaction.debit,
        "credit": Transaction.credit,
        "amountTenge": Transaction.amount_tenge,
    }

    if key == "currency":
        # asc: KZT first, desc: other currencies first
        rank_expr = case((Transaction.currency == "KZT", 0), else_=1)
        if is_desc:
            return [rank_expr.desc(), Transaction.currency.asc(), Transaction.date.desc()]
        return [rank_expr.asc(), Transaction.currency.asc(), Transaction.date.desc()]

    expr = text_exprs.get(key)
    if expr is None:
        expr = numeric_exprs.get(key)
    if expr is None:
        expr = Transaction.date
    primary = expr.desc() if is_desc else expr.asc()
    return [primary, Transaction.date.desc()]


@router.get("", response_model=TransactionListResponse)
async def list_transactions(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    category: Optional[str] = Query(None, description="Transaction category"),
    search: Optional[str] = Query(None, description="Full-text search on purpose"),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Sender IIN/BIN/Account"),
    recipient: Optional[str] = Query(None, description="Recipient IIN/BIN/Account"),
    scope: str = Query("bank", description="bank | all"),
    sort_by: str = Query("date"),
    sort_dir: str = Query("desc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    include_esf = (scope or "bank").strip().lower() == "all"
    order_by = _build_transaction_order_by(sort_by, sort_dir)
    where_clause = _build_transaction_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        category=category,
        search=search,
        min_amount=min_amount,
        max_amount=max_amount,
        currency=currency,
        sender=sender,
        recipient=recipient,
        include_esf=include_esf,
    )

    count_q = select(func.count(Transaction.id)).where(where_clause)
    total = (await db.execute(count_q)).scalar() or 0

    sum_q = select(
        func.coalesce(func.sum(Transaction.debit), 0),
        func.coalesce(func.sum(Transaction.credit), 0),
    ).where(where_clause)
    sums = (await db.execute(sum_q)).one()
    total_debit, total_credit = float(sums[0]), float(sums[1])

    offset = (page - 1) * per_page
    rows_q = (
        select(Transaction, TransactionUploadMeta.uploaded_by_email.label("uploaded_by_email"))
        .outerjoin(TransactionUploadMeta, TransactionUploadMeta.tx_id == Transaction.id)
        .where(where_clause)
        .order_by(*order_by)
        .offset(offset)
        .limit(per_page)
    )
    rows = (await db.execute(rows_q)).all()
    total_pages = max(1, (total + per_page - 1) // per_page)

    data = [
        TransactionOut(
            id=str(t.id),
            date=_format_transaction_dt(t),
            sender=CounterpartyOut(
                name=_fix_mojibake(t.sender_name) or "",
                iin_bin=t.sender_iin_bin or "",
                account=t.sender_account or "",
            ),
            recipient=CounterpartyOut(
                name=_fix_mojibake(t.recipient_name) or "",
                iin_bin=t.recipient_iin_bin or "",
                account=t.recipient_account or "",
            ),
            category=_derive_display_category(
                t.category or "",
                t.purpose or "",
                t.operation_type or "",
                t.direction or "",
                transaction_category=getattr(t, "transaction_category", "") or "",
            ),
            transaction_category=_normalize_display_label(getattr(t, "transaction_category", "") or ""),
            operation_type=_normalize_display_label(t.operation_type) or "",
            purpose=_fix_mojibake(t.purpose) or "",
            currency=t.currency or "",
            debit=float(t.debit or 0),
            credit=float(t.credit or 0),
            amount_tenge=float(t.amount_tenge or 0),
            uploaded_by_email=uploaded_by_email or "",
        )
        for t, uploaded_by_email in rows
    ]

    return TransactionListResponse(
        data=data,
        pagination=PaginationOut(page=page, per_page=per_page, total=total, total_pages=total_pages),
        summary=SummaryOut(total_debit=total_debit, total_credit=total_credit),
    )


@router.get("/export")
async def export_transactions(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    category: Optional[str] = Query(None, description="Transaction category"),
    search: Optional[str] = Query(None, description="Full-text search on purpose"),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Sender IIN/BIN/Account"),
    recipient: Optional[str] = Query(None, description="Recipient IIN/BIN/Account"),
    scope: str = Query("bank", description="bank | all"),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    include_esf = (scope or "bank").strip().lower() == "all"
    where_clause = _build_transaction_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        category=category,
        search=search,
        min_amount=min_amount,
        max_amount=max_amount,
        currency=currency,
        sender=sender,
        recipient=recipient,
        include_esf=include_esf,
    )

    rows_q = (
        select(Transaction, TransactionUploadMeta.uploaded_by_email.label("uploaded_by_email"))
        .outerjoin(TransactionUploadMeta, TransactionUploadMeta.tx_id == Transaction.id)
        .where(where_clause)
        .order_by(Transaction.date.desc())
    )
    rows = (await db.execute(rows_q)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append([
        "Категория",
        "Вид операции",
        "Дата",
        "Отправитель",
        "ИИН/БИН отправителя",
        "Счет отправителя",
        "Получатель",
        "ИИН/БИН получателя",
        "Счет получателя",
        "Назначение",
        "Валюта",
        "Расход",
        "Поступление",
        "Сумма (тенге)",
        "Кто добавил",
    ])

    for t, uploaded_by_email in rows:
        ws.append([
            _derive_display_category(
                t.category or "",
                t.purpose or "",
                t.operation_type or "",
                t.direction or "",
                transaction_category=getattr(t, "transaction_category", "") or "",
            ),
            _normalize_display_label(t.operation_type or ""),
            _format_transaction_dt(t),
            _fix_mojibake(t.sender_name or ""),
            t.sender_iin_bin or "",
            t.sender_account or "",
            _fix_mojibake(t.recipient_name or ""),
            t.recipient_iin_bin or "",
            t.recipient_account or "",
            _fix_mojibake(t.purpose or ""),
            t.currency or "",
            float(t.debit or 0),
            float(t.credit or 0),
            float(t.amount_tenge or 0),
            uploaded_by_email or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now_label = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transactions_{now_label}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _effective_esf_dt(record: EsfRecord) -> Optional[datetime]:
    return record.turnover_date or record.issue_date


def _format_esf_dt(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return value.strftime("%d.%m.%Y %H:%M")


def _build_esf_where_clause(
    project_id: str,
    date: Optional[str],
    search: Optional[str],
    currency: Optional[str],
    sender: Optional[str],
    recipient: Optional[str],
):
    conditions = [EsfRecord.project_id == project_id]
    if date:
        dt = _parse_date(date).date()
        conditions.append(
            or_(
                cast(func.timezone("UTC", EsfRecord.issue_date), Date) == dt,
                cast(func.timezone("UTC", EsfRecord.turnover_date), Date) == dt,
            )
        )
    if search:
        query = search.strip()
        search_conditions = [
            _text_match_conditions(EsfRecord.registration_number, query),
            _text_match_conditions(EsfRecord.tru_name, query),
            _text_match_conditions(EsfRecord.contract_number, query),
        ]
        conditions.append(or_(*[cond for cond in search_conditions if cond is not None]))
    if currency:
        conditions.append(func.upper(func.coalesce(EsfRecord.currency_code, "")) == currency.upper())
    if sender:
        sender_q = sender.strip()
        sender_name_cond = _text_match_conditions(EsfRecord.supplier_name, sender_q)
        conditions.append(
            or_(
                EsfRecord.supplier_iin_bin.ilike(f"%{sender_q}%"),
                sender_name_cond if sender_name_cond is not None else False,
            )
        )
    if recipient:
        recipient_q = recipient.strip()
        buyer_name_cond = _text_match_conditions(EsfRecord.buyer_name, recipient_q)
        conditions.append(
            or_(
                EsfRecord.buyer_iin_bin.ilike(f"%{recipient_q}%"),
                buyer_name_cond if buyer_name_cond is not None else False,
            )
        )
    return and_(*conditions)


def _build_esf_order_by(sort_by: str, sort_dir: str):
    key = (sort_by or "date").strip()
    direction = (sort_dir or "desc").strip().lower()
    is_desc = direction != "asc"

    mapping = {
        "date": func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date),
        "operation_type": func.lower(func.coalesce(EsfRecord.esf_status, "")),
        "sender": func.lower(func.coalesce(EsfRecord.supplier_name, "")),
        "recipient": func.lower(func.coalesce(EsfRecord.buyer_name, "")),
        "supplier_iin_bin": func.lower(func.coalesce(EsfRecord.supplier_iin_bin, "")),
        "supplier_name": func.lower(func.coalesce(EsfRecord.supplier_name, "")),
        "buyer_iin_bin": func.lower(func.coalesce(EsfRecord.buyer_iin_bin, "")),
        "buyer_name": func.lower(func.coalesce(EsfRecord.buyer_name, "")),
        "tru_name": func.lower(func.coalesce(EsfRecord.tru_name, "")),
        "quantity": EsfRecord.quantity,
        "unit": func.lower(func.coalesce(EsfRecord.unit, "")),
        "vat_rate": EsfRecord.vat_rate,
        "price_without_vat": EsfRecord.price_without_vat,
        "price_with_vat": EsfRecord.price_with_vat,
        "purpose": func.lower(func.coalesce(EsfRecord.tru_name, "")),
        "currency": func.lower(func.coalesce(EsfRecord.currency_code, "")),
        "amount_tenge": EsfRecord.total_amount,
        "debit": EsfRecord.total_amount,
        "credit": EsfRecord.total_amount,
        "category": func.lower(func.coalesce(EsfRecord.esf_direction, "")),
    }
    expr = mapping.get(key, mapping["date"])
    primary = expr.desc() if is_desc else expr.asc()
    return [primary, func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date).desc()]


def _is_esf_numeric_key(key: str) -> bool:
    return bool(
        key.startswith("year_")
        or key.startswith("qty_")
        or key.startswith("amt_")
        or key in {
            "overall_total",
            "total_quantity",
            "total_amount",
            "price_with_vat",
            "price_without_vat",
            "vat_rate",
        }
    )


def _append_esf_group_totals(
    rows: list[dict],
    grouped_rows: list[dict],
    group_key: tuple[str, str],
    total_label: str = "ИТОГ",
):
    if not grouped_rows:
        return

    subtotal: dict[str, Any] = {
        "id": hashlib.sha1(f"subtotal|{'|'.join(group_key)}".encode("utf-8", errors="ignore")).hexdigest(),
        "row_type": "subtotal",
        "label": total_label,
        "buyer_iin_bin": "",
        "buyer_name": "",
        "supplier_iin_bin": "",
        "supplier_name": "",
        "tru_name": "",
        "unit": "",
        "currency": "",
    }
    for row in grouped_rows:
        for key, value in row.items():
            if _is_esf_numeric_key(key):
                subtotal[key] = float(subtotal.get(key, 0) or 0) + float(value or 0)

    rows.extend(grouped_rows)
    rows.append(subtotal)


def _append_esf_grand_total(rows: list[dict], total_label: str = "Общий итог"):
    if not rows:
        return

    grand_total: dict[str, Any] = {
        "id": hashlib.sha1(f"grand-total|{len(rows)}".encode("utf-8", errors="ignore")).hexdigest(),
        "row_type": "grand_total",
        "label": total_label,
        "buyer_iin_bin": "",
        "buyer_name": "",
        "supplier_iin_bin": "",
        "supplier_name": "",
        "tru_name": "",
        "unit": "",
        "currency": "",
    }
    for row in rows:
        if row.get("row_type") != "detail":
            continue
        for key, value in row.items():
            if _is_esf_numeric_key(key):
                grand_total[key] = float(grand_total.get(key, 0) or 0) + float(value or 0)

    rows.append(grand_total)


def _build_esf_summary_from_records(records: list[EsfRecord]) -> tuple[list[dict], list[int]]:
    years = sorted(
        {
            int(record.year)
            for record in records
            if record.year is not None
        }
    )
    grouped: dict[tuple[str, str, str, str], dict] = {}

    for record in records:
        key = (
            record.buyer_iin_bin or "",
            _fix_mojibake(record.buyer_name or ""),
            record.supplier_iin_bin or "",
            _fix_mojibake(record.supplier_name or ""),
        )
        year = int(record.year or 0)
        row = grouped.setdefault(
            key,
            {
                "id": hashlib.sha1("|".join(key).encode("utf-8", errors="ignore")).hexdigest(),
                "buyer_iin_bin": key[0],
                "buyer_name": key[1],
                "supplier_iin_bin": key[2],
                "supplier_name": key[3],
                "overall_total": 0.0,
            },
        )
        amount = float(record.total_amount or 0)
        row["overall_total"] += amount
        if year:
            row[f"year_{year}"] = float(row.get(f"year_{year}", 0)) + amount

    rows = []
    for row in grouped.values():
        detail = dict(row)
        detail["row_type"] = "detail"
        rows.append(detail)

    rows.sort(
        key=lambda item: (
            -float(item.get("overall_total", 0) or 0),
            item.get("buyer_name", ""),
            item.get("supplier_name", ""),
        )
    )
    return rows, years


def _build_esf_tru_summary_from_records(records: list[EsfRecord]) -> tuple[list[dict], list[int]]:
    years = sorted(
        {
            int(record.year)
            for record in records
            if record.year is not None
        }
    )
    grouped: dict[tuple[str, str, str, str, str, float, float, float, str, str], dict] = {}

    for record in records:
        key = (
            record.buyer_iin_bin or "",
            _fix_mojibake(record.buyer_name or ""),
            record.supplier_iin_bin or "",
            _fix_mojibake(record.supplier_name or ""),
            _fix_mojibake(record.tru_name or ""),
            float(record.price_with_vat or 0),
            float(record.price_without_vat or 0),
            float(record.vat_rate or 0),
            _fix_mojibake(record.unit or ""),
            (record.currency_code or "").upper(),
        )
        year = int(record.year or 0)
        row = grouped.setdefault(
            key,
            {
                "id": hashlib.sha1("|".join(str(part) for part in key).encode("utf-8", errors="ignore")).hexdigest(),
                "buyer_iin_bin": key[0],
                "buyer_name": key[1],
                "supplier_iin_bin": key[2],
                "supplier_name": key[3],
                "tru_name": key[4],
                "price_with_vat": key[5],
                "price_without_vat": key[6],
                "vat_rate": key[7],
                "unit": key[8],
                "currency": key[9],
                "total_quantity": 0.0,
                "total_amount": 0.0,
            },
        )
        amount = float(record.total_amount or 0)
        qty = float(record.quantity or 0)
        row["total_quantity"] += qty
        row["total_amount"] += amount
        if year:
            row[f"qty_{year}"] = float(row.get(f"qty_{year}", 0)) + qty
            row[f"amt_{year}"] = float(row.get(f"amt_{year}", 0)) + amount

    detail_rows = []
    for row in grouped.values():
        detail = dict(row)
        detail["row_type"] = "detail"
        detail_rows.append(detail)

    grouped_by_buyer: dict[tuple[str, str], list[dict]] = {}
    for row in detail_rows:
        buyer_key = (row.get("buyer_iin_bin", ""), row.get("buyer_name", ""))
        grouped_by_buyer.setdefault(buyer_key, []).append(row)

    buyer_groups = sorted(
        grouped_by_buyer.items(),
        key=lambda item: sum(float(row.get("total_amount", 0) or 0) for row in item[1]),
        reverse=True,
    )

    rows: list[dict] = []
    for buyer_key, buyer_rows in buyer_groups:
        buyer_rows.sort(
            key=lambda item: (
                -float(item.get("total_amount", 0) or 0),
                item.get("tru_name", ""),
            )
        )
        _append_esf_group_totals(rows, buyer_rows, buyer_key)

    _append_esf_grand_total(rows)
    return rows, years


def _paginate_rows(rows: list[dict], page: int, per_page: int) -> tuple[list[dict], int]:
    total = len(rows)
    start = max(0, (page - 1) * per_page)
    end = start + per_page
    return rows[start:end], total


@router.get("/esf", response_model=EsfListResponse)
async def list_esf_records(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    search: Optional[str] = Query(None, description="Search by reg number / TRU / contract"),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Supplier IIN/BIN/Name"),
    recipient: Optional[str] = Query(None, description="Buyer IIN/BIN/Name"),
    sort_by: str = Query("date"),
    sort_dir: str = Query("desc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    where_clause = _build_esf_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        search=search,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )

    total = (await db.execute(select(func.count(EsfRecord.id)).where(where_clause))).scalar() or 0
    sums = (
        await db.execute(
            select(
                func.coalesce(func.sum(EsfRecord.total_amount), 0),
                func.coalesce(func.sum(EsfRecord.vat_amount), 0),
                func.coalesce(func.sum(case((EsfRecord.esf_direction == "sale", EsfRecord.total_amount), else_=0)), 0),
                func.coalesce(func.sum(case((EsfRecord.esf_direction == "purchase", EsfRecord.total_amount), else_=0)), 0),
            ).where(where_clause)
        )
    ).one()
    total_amount, total_vat, sales_amount, purchase_amount = [float(value or 0) for value in sums]

    rows = (
        await db.execute(
            select(EsfRecord)
            .where(where_clause)
            .order_by(*_build_esf_order_by(sort_by, sort_dir))
            .offset((page - 1) * per_page)
            .limit(per_page)
        )
    ).scalars().all()

    total_pages = max(1, (total + per_page - 1) // per_page)
    data = [
        EsfRecordOut(
            id=str(record.id),
            registration_number=record.registration_number,
            esf_status=record.esf_status or "",
            esf_direction=record.esf_direction or "",
            issue_date=_format_esf_dt(record.issue_date),
            turnover_date=_format_esf_dt(record.turnover_date),
            supplier=EsfCounterpartyOut(
                name=_fix_mojibake(record.supplier_name or ""),
                iin_bin=record.supplier_iin_bin or "",
                address=_fix_mojibake(record.supplier_address or ""),
            ),
            buyer=EsfCounterpartyOut(
                name=_fix_mojibake(record.buyer_name or ""),
                iin_bin=record.buyer_iin_bin or "",
                address=_fix_mojibake(record.buyer_address or ""),
            ),
            tru_name=_fix_mojibake(record.tru_name or ""),
            quantity=float(record.quantity or 0),
            unit=_fix_mojibake(record.unit or ""),
            vat_rate=float(record.vat_rate or 0),
            price_without_vat=float(record.price_without_vat or 0),
            price_with_vat=float(record.price_with_vat or 0),
            total_amount=float(record.total_amount or 0),
            vat_amount=float(record.vat_amount or 0),
            currency_code=record.currency_code or "",
            contract_number=_fix_mojibake(record.contract_number or ""),
        )
        for record in rows
    ]

    return EsfListResponse(
        data=data,
        pagination=PaginationOut(page=page, per_page=per_page, total=total, total_pages=total_pages),
        summary=EsfSummaryOut(
            total_records=total,
            total_amount=total_amount,
            total_vat=total_vat,
            sales_amount=sales_amount,
            purchase_amount=purchase_amount,
        ),
    )


@router.get("/esf/summary", response_model=EsfSheetResponse)
async def list_esf_summary_sheet(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    search: Optional[str] = Query(None, description="Search by reg number / TRU / contract"),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Supplier IIN/BIN/Name"),
    recipient: Optional[str] = Query(None, description="Buyer IIN/BIN/Name"),
    direction: str = Query("sale", description="sale | purchase"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    where_clause = _build_esf_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        search=search,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )
    records = (
        await db.execute(
            select(EsfRecord)
            .where(where_clause)
            .order_by(func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date).desc())
        )
    ).scalars().all()
    rows, years = _build_esf_summary_from_records(records)
    page_rows, total = _paginate_rows(rows, page, per_page)
    total_amount = float(sum(float(record.total_amount or 0) for record in records))
    total_vat = float(sum(float(record.vat_amount or 0) for record in records))
    sales_amount = total_amount
    purchase_amount = total_amount

    return EsfSheetResponse(
        data=page_rows,
        pagination=PaginationOut(page=page, per_page=per_page, total=total, total_pages=max(1, (total + per_page - 1) // per_page)),
        summary=EsfSummaryOut(
            total_records=total,
            total_amount=total_amount,
            total_vat=total_vat,
            sales_amount=sales_amount,
            purchase_amount=purchase_amount,
        ),
        years=years,
    )


@router.get("/esf/tru-summary", response_model=EsfSheetResponse)
async def list_esf_tru_summary_sheet(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    search: Optional[str] = Query(None, description="Search by reg number / TRU / contract"),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Supplier IIN/BIN/Name"),
    recipient: Optional[str] = Query(None, description="Buyer IIN/BIN/Name"),
    direction: str = Query("sale", description="sale | purchase"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    where_clause = _build_esf_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        search=search,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )
    records = (
        await db.execute(
            select(EsfRecord)
            .where(where_clause)
            .order_by(func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date).desc())
        )
    ).scalars().all()
    rows, years = _build_esf_tru_summary_from_records(records)
    page_rows, total = _paginate_rows(rows, page, per_page)
    total_amount = float(sum(float(record.total_amount or 0) for record in records))
    total_vat = float(sum(float(record.vat_amount or 0) for record in records))
    sales_amount = total_amount
    purchase_amount = total_amount

    return EsfSheetResponse(
        data=page_rows,
        pagination=PaginationOut(page=page, per_page=per_page, total=total, total_pages=max(1, (total + per_page - 1) // per_page)),
        summary=EsfSummaryOut(
            total_records=total,
            total_amount=total_amount,
            total_vat=total_vat,
            sales_amount=sales_amount,
            purchase_amount=purchase_amount,
        ),
        years=years,
    )


@router.get("/esf/export")
async def export_esf_records(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    search: Optional[str] = Query(None, description="Search by reg number / TRU / contract"),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Supplier IIN/BIN/Name"),
    recipient: Optional[str] = Query(None, description="Buyer IIN/BIN/Name"),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    where_clause = _build_esf_where_clause(
        project_id=ctx.project.project_id,
        date=date,
        search=search,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )
    rows = (
        await db.execute(
            select(EsfRecord)
            .where(where_clause)
            .order_by(func.coalesce(EsfRecord.turnover_date, EsfRecord.issue_date).desc())
        )
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "ESF"
    ws.append([
        "Регистрационный номер ЭСФ",
        "Статус ЭСФ",
        "Направление",
        "Дата выписки",
        "Дата совершения оборота",
        "ИИН/БИН поставщика",
        "Поставщик",
        "ИИН/БИН покупателя",
        "Покупатель",
        "Наименование ТРУ",
        "Номер договора",
        "Общая сумма",
        "Сумма НДС",
        "Код валюты",
    ])
    for record in rows:
        ws.append([
            record.registration_number,
            record.esf_status or "",
            record.esf_direction or "",
            _format_esf_dt(record.issue_date),
            _format_esf_dt(record.turnover_date),
            record.supplier_iin_bin or "",
            _fix_mojibake(record.supplier_name or ""),
            record.buyer_iin_bin or "",
            _fix_mojibake(record.buyer_name or ""),
            _fix_mojibake(record.tru_name or ""),
            _fix_mojibake(record.contract_number or ""),
            float(record.total_amount or 0),
            float(record.vat_amount or 0),
            record.currency_code or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    now_label = datetime.now().strftime("%Y%m%d_%H%M%S")
    headers = {"Content-Disposition": f'attachment; filename="esf_{now_label}.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import-statement", response_model=TransactionImportResponse)
async def import_statement(
    file: UploadFile = File(...),
    parser_type: str = Form(PARSER_KASPI),
    ctx: ProjectContext = Depends(get_current_project_context),
    db: AsyncSession = Depends(get_db),
):
    uploader_email = str(ctx.user.email or "").strip().lower()
    parser_type = (parser_type or PARSER_KASPI).strip().lower()
    project_id = ctx.project.project_id

    filename = (file.filename or "").lower()
    allowed_ext = (".xls", ".xlsx")
    if parser_type == PARSER_TRANSACTIONS_CORE:
        allowed_ext = (".csv",)
    elif parser_type == PARSER_ESF:
        allowed_ext = (".csv", ".xls", ".xlsx")

    if not filename.endswith(allowed_ext):
        if parser_type == PARSER_TRANSACTIONS_CORE:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .csv for transactions_core")
        if parser_type == PARSER_ESF:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .csv, .xls or .xlsx for esf")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .xls or .xlsx")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")

    if parser_type == PARSER_TRANSACTIONS_CORE:
        tx_rows, skipped = _extract_transactions_from_transactions_core_csv(content)
        file_uuid = str(uuid4())
        source_sheet = file.filename or "transactions_core.csv"
        await db.execute(
            text(
                """
                INSERT INTO afm.raw_files(file_id, project_id, source_bank, original_filename, sha256, parser_version)
                VALUES (CAST(:file_id AS uuid), CAST(:project_id AS uuid), :source_bank, :filename, :sha256, :parser_version)
                ON CONFLICT (file_id) DO NOTHING
                """
            ),
            {
                "file_id": file_uuid,
                "project_id": project_id,
                "source_bank": "transactions_core",
                "filename": source_sheet,
                "sha256": hashlib.sha256(content).hexdigest(),
                "parser_version": settings.parser_version,
            },
        )
        payloads = []
        for idx, tx in enumerate(tx_rows, start=2):
            payloads.append(
                _build_core_transaction_payload(
                    tx,
                    uploader_email=uploader_email,
                    project_id=project_id,
                    source_bank="transactions_core",
                    source_sheet=source_sheet,
                    source_row_no=idx,
                    file_id=file_uuid,
                    statement_id=None,
                    format_id=None,
                    raw_row_json=tx,
                )
            )

        unique_payloads = []
        seen_hashes = set()
        for payload in payloads:
            row_hash = payload["row_hash"]
            if row_hash in seen_hashes:
                skipped += 1
                continue
            seen_hashes.add(row_hash)
            unique_payloads.append(payload)

        existing_hashes = set(
            (
                await db.execute(
                    select(Transaction.row_hash).where(
                        Transaction.project_id == project_id,
                        Transaction.row_hash.in_([payload["row_hash"] for payload in unique_payloads]),
                    )
                )
            ).scalars()
        )

        inserted = 0
        for payload in unique_payloads:
            if payload["row_hash"] in existing_hashes:
                skipped += 1
                continue

            db.add(Transaction(**payload))
            db.add(
                TransactionUploadMeta(
                    tx_id=payload["id"],
                    project_id=project_id,
                    uploaded_by_email=uploader_email or "",
                    created_at=datetime.utcnow(),
                )
            )
            inserted += 1

        await db.commit()
        fraud_warnings = await _build_import_fraud_warnings_for_file(file_uuid, project_id, db)
        fraud_warnings.extend(await _build_project_reconciliation_warnings(project_id, db))
        return TransactionImportResponse(
            inserted=inserted,
            skipped=skipped,
            message=f"Imported {inserted} transactions from transactions_core CSV",
            fraud_warnings=fraud_warnings,
        )
    if parser_type == PARSER_ESF:
        esf_rows, owner_side, skipped = _extract_esf_records(content, file.filename or "")
        file_uuid = str(uuid4())
        source_sheet = file.filename or "esf"
        await db.execute(
            text(
                """
                INSERT INTO afm.raw_files(file_id, project_id, source_bank, original_filename, sha256, parser_version)
                VALUES (CAST(:file_id AS uuid), CAST(:project_id AS uuid), :source_bank, :filename, :sha256, :parser_version)
                ON CONFLICT (file_id) DO NOTHING
                """
            ),
            {
                "file_id": file_uuid,
                "project_id": project_id,
                "source_bank": "esf",
                "filename": source_sheet,
                "sha256": hashlib.sha256(content).hexdigest(),
                "parser_version": settings.parser_version,
            },
        )

        payloads: list[dict] = []
        for idx, row in enumerate(esf_rows, start=2):
            payload = _build_esf_record_payload(
                row,
                owner_side=owner_side,
                project_id=project_id,
                file_id=file_uuid,
                source_sheet=source_sheet,
                source_row_no=idx,
            )
            if payload is None:
                skipped += 1
                continue
            payloads.append(payload)

        unique_payloads: list[dict] = []
        seen_hashes = set()
        for payload in payloads:
            row_hash = payload["row_hash"]
            if row_hash in seen_hashes:
                skipped += 1
                continue
            seen_hashes.add(row_hash)
            unique_payloads.append(payload)

        existing_hashes = set(
            (
                await db.execute(
                    select(EsfRecord.row_hash).where(
                        EsfRecord.project_id == project_id,
                        EsfRecord.row_hash.in_([payload["row_hash"] for payload in unique_payloads]),
                    )
                )
            ).scalars()
        )

        inserted = 0
        for payload in unique_payloads:
            if payload["row_hash"] in existing_hashes:
                skipped += 1
                continue

            db.add(EsfRecord(**payload))
            shadow_payload = _build_esf_shadow_transaction_payload(
                payload,
                uploader_email=uploader_email,
            )
            db.add(Transaction(**shadow_payload))
            db.add(
                TransactionUploadMeta(
                    tx_id=shadow_payload["tx_id"],
                    project_id=project_id,
                    uploaded_by_email=uploader_email or "",
                    created_at=datetime.utcnow(),
                )
            )
            inserted += 1

        await db.commit()
        reconciliation_warnings = await _build_project_reconciliation_warnings(project_id, db)
        return TransactionImportResponse(
            inserted=inserted,
            skipped=skipped,
            message=f"Imported {inserted} ESF rows",
            fraud_warnings=reconciliation_warnings,
        )
    source_bank = _parser_type_to_source_bank(parser_type)

    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # The sync ingestion pipeline needs its own psycopg2 connection. Release
        # the request-scoped async connection first so tiny Postgres instances
        # don't fail with "remaining connection slots are reserved".
        await db.close()
        await async_engine.dispose()
        result = await run_in_threadpool(_run_ingestion_pipeline, tmp_path, source_bank, project_id)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Ingestion pipeline failed: {exc}") from exc
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    parsed_rows = int(result.get("core_rows") or 0)
    inserted = await _finalize_ingestion_import(result.get("file_id"), uploader_email, project_id)
    skipped = max(parsed_rows - inserted, 0)
    fraud_warnings = await _build_import_fraud_warnings_for_file(result.get("file_id"), project_id)
    fraud_warnings.extend(await _build_project_reconciliation_warnings(project_id))

    bank_label = result.get("bank") or source_bank or "auto"
    return TransactionImportResponse(
        inserted=inserted,
        skipped=skipped,
        message=f"Imported {inserted} transactions via {bank_label}",
        fraud_warnings=fraud_warnings,
    )


