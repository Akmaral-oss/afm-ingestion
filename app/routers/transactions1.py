"""
Transactions API:
- GET /api/v1/transactions
- POST /api/v1/transactions/import-statement
"""

import csv
import hashlib
import json
import os
from datetime import datetime
from io import BytesIO, StringIO
from tempfile import NamedTemporaryFile
from typing import Optional
import inspect
import re
from uuid import uuid4

import httpx
from fastapi.concurrency import run_in_threadpool
from fastapi import APIRouter, Depends, Query, File, UploadFile, Header, HTTPException, status, Form
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
from sqlalchemy import select, func, and_, or_, case
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..database import get_db, async_session, engine as async_engine
from ..ingestion.pipeline import IngestionPipeline
from ..models import Transaction, TransactionUploadMeta
from ..schemas import (
    TransactionListResponse,
    TransactionOut,
    CounterpartyOut,
    PaginationOut,
    SummaryOut,
    TransactionImportResponse,
)
from ..security import decode_access_token

router = APIRouter(prefix="/transactions", tags=["Transactions"])

PARSER_BANK = "bank_parser"
PARSER_KASPI = "kaspi"
PARSER_KASPI_LEGACY = "kaspi_parser"
PARSER_HALYK = "halyk_parser"
PARSER_SMART = "smart_parser"
PARSER_TRANSACTIONS_CORE = "transactions_core"


def _parse_date(date_str: str) -> datetime:
    return datetime.strptime(date_str, "%d.%m.%Y")


def _effective_transaction_dt(tx: Transaction) -> Optional[datetime]:
    if tx.date is not None:
        return tx.date
    if tx.operation_date is not None:
        return datetime.combine(tx.operation_date, datetime.min.time())
    return None


def _format_transaction_dt(tx: Transaction) -> str:
    dt = _effective_transaction_dt(tx)
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""


def _parse_operation_datetime(raw: object) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw

    s = str(raw).strip()
    if not s:
        return None

    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _decode_text_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return 0.0

    s = s.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _safe_amount_tenge(amount_kzt: float, debit: float, credit: float) -> float:
    """
    Protect against parser misalignment where amount_kzt may accidentally contain
    huge header numbers (e.g., scientific-notation IIN from statement header).
    """
    base = max(float(debit or 0), float(credit or 0))
    amount = float(amount_kzt or 0)

    if amount <= 0:
        return base
    if base <= 0:
        return amount

    # If amount is far larger than transaction side amount, treat it as corrupted.
    ratio = amount / base if base > 0 else 1.0
    if ratio > 1000 or amount > 1e10:
        return base
    return amount


def _to_str(value: object, max_len: int = 255) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s in {"'", '"'}:
        return ""
    return s[:max_len]


def _fix_mojibake(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        # Fix strings like "РїРµСЂРµРІРѕРґ" -> "перевод".
        fixed = s.encode("cp1251").decode("utf-8")
        if fixed:
            s = fixed
    except Exception:
        pass
    # Common parser-side artifact for "ё".
    s = s.replace("С‘", "ё").replace("Рµ", "е")
    return s


def _derive_display_category(category: str, purpose_text: str, operation_type_raw: str, direction: str = "") -> str:
    base = _fix_mojibake(category or "").strip()
    if base:
        return base
    return _derive_category_from_core_row(base, purpose_text, operation_type_raw, direction)


def _derived_category_expr():
    trimmed_category = func.nullif(func.trim(Transaction.category), "")
    purpose = func.lower(func.coalesce(Transaction.purpose, ""))
    op_type = func.lower(func.coalesce(Transaction.operation_type, ""))
    direction = func.lower(func.coalesce(Transaction.direction, ""))
    return func.coalesce(
        trimmed_category,
        case(
            (or_(purpose.like("%red.kz%"), purpose.like("%продаж%")), "Продажи Kaspi / Red"),
            (or_(purpose.like("%погаш%"), purpose.like("%кредит%")), "Погашение кредита"),
            (or_(purpose.like("%снятие%"), purpose.like("%cash%"), purpose.like("%atm%")), "Снятие наличных"),
            (or_(purpose.like("%пополн%"), purpose.like("%взнос%"), purpose.like("%deposit%")), "Пополнение счёта"),
            (purpose.like("%рассроч%"), "Рассрочка Kaspi"),
            (purpose.like("%займ%"), "Выдача займа"),
            (and_(purpose.like("%перевод%"), purpose.like("%внутр%")), "Внутренние операции"),
            (or_(purpose.like("%перевод%"), op_type.like("%payment%")), "P2P перевод"),
            (purpose.like("%оплат%"), "Оплата услуг"),
            (direction == "credit", "Поступления"),
            (direction == "debit", "Расходы"),
            else_="Прочее",
        ),
    )


def _normalize_counterparty_identity(name: str, iin: str, account: str) -> tuple[str, str, str]:
    fixed_name = _fix_mojibake(name or "").strip()
    fixed_iin = re.sub(r"\D+", "", str(iin or ""))[:12]
    fixed_account = (str(account or "").strip().upper())[:20]

    unknown_tokens = {"", "UNKNOWN", "NULL", "NONE", "N/A", "-", "—", "000000000000"}
    if fixed_name.upper() in unknown_tokens:
        fixed_name = ""

    if not fixed_name:
        fixed_name = fixed_account or fixed_iin or ""

    return fixed_name, fixed_iin, fixed_account


def _derive_category_from_core_row(sdp_name: str, purpose_text: str, operation_type_raw: str, direction: str) -> str:
    sdp = _fix_mojibake(sdp_name or "").strip()
    if sdp:
        return sdp

    purpose = _fix_mojibake(purpose_text or "").lower()
    op_type = _fix_mojibake(operation_type_raw or "").lower()
    direction_l = (direction or "").strip().lower()

    if "red.kz" in purpose or "продаж" in purpose:
        return "Продажи Kaspi / Red"
    if "погаш" in purpose or "кредит" in purpose:
        return "Погашение кредита"
    if "снятие" in purpose or "cash" in purpose or "atm" in purpose:
        return "Снятие наличных"
    if "пополн" in purpose or "взнос" in purpose or "deposit" in purpose:
        return "Пополнение счёта"
    if "рассроч" in purpose:
        return "Рассрочка Kaspi"
    if "займ" in purpose:
        return "Выдача займа"
    if "перевод" in purpose and "внутр" in purpose:
        return "Внутренние операции"
    if "перевод" in purpose or "payment" in op_type:
        return "P2P перевод"
    if "оплат" in purpose:
        return "Оплата услуг"
    if direction_l == "credit":
        return "Поступления"
    if direction_l == "debit":
        return "Расходы"
    return "Прочее"


def _extract_transactions_from_transactions_core_csv(content: bytes) -> tuple[list[dict], int]:
    text = _decode_text_bytes(content)
    reader = csv.DictReader(StringIO(text))

    out: list[dict] = []
    skipped = 0

    for row in reader:
        if not row:
            continue

        op_dt = _parse_operation_datetime(row.get("operation_ts") or row.get("operation_date"))
        if op_dt is None:
            skipped += 1
            continue

        direction = (row.get("direction") or "").strip().lower()
        amount_currency = _to_float(row.get("amount_currency"))
        debit = _to_float(row.get("amount_debit"))
        credit = _to_float(row.get("amount_credit"))

        if debit <= 0 and credit <= 0 and amount_currency > 0:
            if direction == "debit":
                debit = amount_currency
            elif direction == "credit":
                credit = amount_currency

        amount_tenge = _safe_amount_tenge(_to_float(row.get("amount_kzt")), debit, credit)
        if amount_tenge <= 0:
            skipped += 1
            continue

        sender_name, sender_iin, sender_account = _normalize_counterparty_identity(
            row.get("payer_name") or "",
            row.get("payer_iin_bin") or "",
            row.get("payer_account") or "",
        )
        recipient_name, recipient_iin, recipient_account = _normalize_counterparty_identity(
            row.get("receiver_name") or "",
            row.get("receiver_iin_bin") or "",
            row.get("receiver_account") or "",
        )

        purpose = _fix_mojibake(row.get("purpose_text") or row.get("raw_note") or "").strip()
        operation_type = _fix_mojibake(row.get("operation_type_raw") or "").strip()
        category = _derive_category_from_core_row(
            row.get("sdp_name") or "",
            purpose,
            operation_type,
            direction,
        )

        out.append({
            "date": op_dt,
            "sender_name": sender_name,
            "sender_iin_bin": sender_iin,
            "sender_account": sender_account,
            "recipient_name": recipient_name,
            "recipient_iin_bin": recipient_iin,
            "recipient_account": recipient_account,
            "purpose": purpose,
            "category": category,
            "operation_type": operation_type or direction,
            "currency": _to_str(row.get("currency"), 3).upper() or "KZT",
            "debit": debit,
            "credit": credit,
            "amount_tenge": amount_tenge,
        })

    return out, skipped


def _to_mojibake(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    try:
        return s.encode("utf-8").decode("cp1251")
    except Exception:
        return s


def _norm_header(value: object) -> str:
    s = _fix_mojibake(str(value or "")).strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def _pick_index(index_map: dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        candidates = (
            _norm_header(alias),
            _norm_header(_fix_mojibake(alias)),
            _norm_header(_to_mojibake(alias)),
        )
        for candidate in candidates:
            if candidate in index_map:
                return index_map[candidate]
    return None


def _get_cell(row: tuple, idx: Optional[int]) -> object:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _extract_sheet_client_fallback(rows: list[tuple]) -> tuple[str, str, str]:
    """
    Try to read owner identity from statement header blocks:
    - Клиент:
    - ИИН/БИН:
    - Счет:
    """
    client_name = ""
    client_iin = ""
    client_account = ""

    def _clean_iin(value: str) -> str:
        s = str(value or "").strip().replace("\u00a0", "").replace(" ", "")
        if not s:
            return ""
        sci = s.replace(",", ".")
        if re.search(r"[eE][+-]?\d+", sci):
            try:
                return str(int(round(float(sci))))
            except Exception:
                pass
        return "".join(ch for ch in s if ch.isdigit())

    scan_rows = rows[:40]
    for row in scan_rows:
        if not row:
            continue
        label = _fix_mojibake(_to_str(row[0], 128)).strip().lower().replace("ё", "е")
        if not label:
            continue

        # Usually value is in column B/C in raw exports.
        value = ""
        for idx in (1, 2, 3):
            if idx < len(row):
                cell_val = _fix_mojibake(_to_str(row[idx], 255)).strip()
                if cell_val:
                    value = cell_val
                    break
        if not value:
            continue

        if "клиент" in label and not client_name:
            client_name = value.strip('"').strip("'")
        elif ("иин/бин" in label or "иин" in label or "бин" in label) and not client_iin:
            client_iin = _clean_iin(value)[:12]
        elif ("счет" in label or "счет:" in label or "счета" in label) and not client_account:
            client_account = value

    return client_name, client_iin, client_account


def _is_incoming(direction_value: str, credit: float) -> bool:
    d = _fix_mojibake((direction_value or "").strip()).lower()
    if "вход" in d or "incoming" in d:
        return True
    if "исход" in d or "outgoing" in d:
        return False
    return credit > 0


def _get_parser_url(parser_type: str) -> str:
    parser_type = (parser_type or PARSER_KASPI).strip().lower()
    if parser_type == PARSER_HALYK:
        base = (settings.HALYK_PARSER_API_URL or "").strip()
        if not base:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="HALYK_PARSER_API_URL is not configured",
            )
        return base.rstrip("/") + "/parse"

    if parser_type == PARSER_KASPI:
        base = (
            (settings.KASPI_PARSER_API_URL or "").strip()
            or (settings.BANK_PARSER_API_URL or "").strip()
            or settings.PARSER_API_URL
        )
        return base.rstrip("/") + "/parse"

    if parser_type == PARSER_KASPI_LEGACY:
        base = (
            (settings.BANK_PARSER_API_URL or "").strip()
            or settings.PARSER_API_URL
        )
        return base.rstrip("/")

    if parser_type != PARSER_BANK:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unknown parser_type. Use smart_parser, kaspi, kaspi_parser, bank_parser, halyk_parser or transactions_core",
        )

    base = (settings.BANK_PARSER_API_URL or "").strip() or settings.PARSER_API_URL
    return base.rstrip("/") + "/parse"


def _identity_quality(name: str, iin: str, account: str) -> int:
    score = 0
    n = (name or "").strip().upper()
    i = (iin or "").strip()
    a = (account or "").strip()
    if n and n != "UNKNOWN":
        score += 2
    if i and i not in {"0", "000000000000"}:
        score += 2
    if a:
        score += 1
    return score


def _normalize_id_value(value: str) -> str:
    return "".join(ch for ch in (value or "").strip().upper() if ch.isalnum())


def _normalize_text_key(value: str) -> str:
    return re.sub(r"\s+", " ", _fix_mojibake(value or "").strip().lower())


def _tx_quality(tx: dict) -> int:
    score = 0
    score += _identity_quality(tx.get("sender_name", ""), tx.get("sender_iin_bin", ""), tx.get("sender_account", ""))
    score += _identity_quality(tx.get("recipient_name", ""), tx.get("recipient_iin_bin", ""), tx.get("recipient_account", ""))
    if (tx.get("purpose") or "").strip():
        score += 1
    if (tx.get("category") or "").strip():
        score += 1
    if (tx.get("operation_type") or "").strip():
        score += 1
    return score


def _tx_signature(tx: dict) -> tuple:
    dt: datetime = tx["date"]
    return (
        dt.isoformat(timespec="seconds"),
        (tx.get("currency") or "KZT").strip().upper(),
        round(float(tx.get("debit") or 0), 2),
        round(float(tx.get("credit") or 0), 2),
        round(float(tx.get("amount_tenge") or 0), 2),
        _normalize_id_value(tx.get("sender_iin_bin", "")),
        _normalize_id_value(tx.get("sender_account", "")),
        _normalize_text_key(tx.get("sender_name", ""))[:120],
        _normalize_id_value(tx.get("recipient_iin_bin", "")),
        _normalize_id_value(tx.get("recipient_account", "")),
        _normalize_text_key(tx.get("recipient_name", ""))[:120],
        _normalize_text_key(tx.get("purpose", ""))[:400],
        _normalize_text_key(tx.get("category", ""))[:120],
        _normalize_text_key(tx.get("operation_type", ""))[:120],
    )


def _build_core_transaction_payload(
    tx: dict,
    *,
    uploader_email: str,
    source_bank: str,
    source_sheet: str,
    source_row_no: int,
    file_id: Optional[str] = None,
    statement_id: Optional[str] = None,
    format_id: Optional[str] = None,
    source_block_id: int = 1,
    raw_row_json: Optional[dict] = None,
) -> dict:
    def _jsonable(value):
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, dict):
            return {str(k): _jsonable(v) for k, v in value.items()}
        if isinstance(value, (list, tuple)):
            return [_jsonable(v) for v in value]
        return value

    op_dt: datetime = tx["date"]
    payload = {
        "id": str(tx.get("id") or uuid4()),
        "file_id": file_id,
        "statement_id": statement_id,
        "format_id": format_id,
        "source_bank": source_bank,
        "source_sheet": (source_sheet or "")[:255] or None,
        "source_block_id": source_block_id,
        "source_row_no": source_row_no,
        "date": op_dt,
        "operation_date": op_dt.date(),
        "currency": (tx.get("currency") or "KZT").strip().upper(),
        "amount_currency": float(max(float(tx.get("debit") or 0), float(tx.get("credit") or 0))),
        "amount_tenge": float(tx.get("amount_tenge") or 0),
        "credit": float(tx.get("credit") or 0),
        "debit": float(tx.get("debit") or 0),
        "direction": "credit" if float(tx.get("credit") or 0) > 0 else "debit",
        "operation_type": _fix_mojibake(tx.get("operation_type") or ""),
        "category": _fix_mojibake(tx.get("category") or ""),
        "purpose_code": None,
        "purpose": _fix_mojibake(tx.get("purpose") or ""),
        "raw_note": _fix_mojibake(tx.get("purpose") or ""),
        "sender_name": _fix_mojibake(tx.get("sender_name") or ""),
        "sender_iin_bin": (tx.get("sender_iin_bin") or "")[:32],
        "payer_residency": None,
        "payer_bank": None,
        "sender_account": (tx.get("sender_account") or "")[:64],
        "recipient_name": _fix_mojibake(tx.get("recipient_name") or ""),
        "recipient_iin_bin": (tx.get("recipient_iin_bin") or "")[:32],
        "receiver_residency": None,
        "receiver_bank": None,
        "recipient_account": (tx.get("recipient_account") or "")[:64],
        "confidence_score": 1.0,
        "parse_warnings": None,
        "raw_row_json": _jsonable(raw_row_json) if raw_row_json is not None else None,
    }

    sig_source = json.dumps(
        {
            "date": op_dt.isoformat(timespec="seconds"),
            "currency": payload["currency"],
            "debit": payload["debit"],
            "credit": payload["credit"],
            "sender_iin_bin": payload["sender_iin_bin"],
            "sender_account": payload["sender_account"],
            "recipient_iin_bin": payload["recipient_iin_bin"],
            "recipient_account": payload["recipient_account"],
            "purpose": payload["purpose"],
            "category": payload["category"],
            "operation_type": payload["operation_type"],
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    payload["row_hash"] = hashlib.sha256(sig_source.encode("utf-8")).hexdigest()
    return payload


def _make_xlsx_with_single_sheet(sheet_title: str, rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    for row in rows:
        ws.append(list(row))
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _extract_transactions_from_workbook(workbook, parser_type: str) -> tuple[list[dict], int]:
    parser_type = (parser_type or "").strip().lower()
    out: list[dict] = []
    skipped = 0

    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        sheet_client_name, sheet_client_iin, sheet_client_account = _extract_sheet_client_fallback(rows)

        header_norm = [_norm_header(c) for c in rows[0]]
        index = {name: i for i, name in enumerate(header_norm)}

        if parser_type in {PARSER_HALYK, PARSER_KASPI}:
            idx_date = _pick_index(index, "РґР°С‚Р° РѕРїРµСЂР°С†РёРё", "operation_date")
            idx_currency = _pick_index(index, "РІР°Р»СЋС‚Р° РѕРїРµСЂР°С†РёРё", "РІР°Р»СЋС‚Р°", "currency")
            idx_debit = _pick_index(index, "СЃСѓРјРјР° РґРµР±РµС‚", "РґРµР±РµС‚", "debit")
            idx_credit = _pick_index(index, "СЃСѓРјРјР° РєСЂРµРґРёС‚", "РєСЂРµРґРёС‚", "credit")
            idx_amount = _pick_index(index, "СЃСѓРјРјР° РІ С‚РµРЅРіРµ", "amount_kzt")
            idx_direction = _pick_index(index, "РЅР°РїСЂР°РІР»РµРЅРёРµ", "direction")
            idx_owner_name = _pick_index(index, "РІР»Р°РґРµР»РµС† СЃС‡РµС‚Р°", "owner_name")
            idx_owner_iin = _pick_index(index, "РёРёРЅ/Р±РёРЅ РІР»Р°РґРµР»СЊС†Р°", "owner_iin")
            idx_owner_acc = _pick_index(index, "РЅРѕРјРµСЂ СЃС‡РµС‚Р° РІР»Р°РґРµР»СЊС†Р°", "РЅРѕРјРµСЂ СЃС‡РµС‚Р° РІР»Р°РґРµР»СЊС†Р°", "РЅРѕРјРµСЂ СЃС‡С‘С‚Р° РІР»Р°РґРµР»СЊС†Р°", "owner_account")
            idx_cp_name = _pick_index(index, "РєРѕРЅС‚СЂР°РіРµРЅС‚", "recipient", "counterparty")
            idx_cp_iin = _pick_index(index, "РёРёРЅ/Р±РёРЅ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "iin_recipient", "counterparty_iin")
            idx_cp_acc = _pick_index(index, "СЃС‡РµС‚ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "СЃС‡С‘С‚ РєРѕРЅС‚СЂР°РіРµРЅС‚Р°", "account_recipient", "counterparty_account")
            idx_purpose = _pick_index(index, "РЅР°Р·РЅР°С‡РµРЅРёРµ РїР»Р°С‚РµР¶Р°", "transfer_purpose", "purpose")
            idx_category = _pick_index(index, "РєР°С‚РµРіРѕСЂРёСЏ", "category")
            idx_operation_type = _pick_index(
                index,
                "operation_type",
                "document_category",
                "op_type",
                "type",
                "РўРёРї РѕРїРµСЂР°С†РёРё",
                "РІРёРґС‹ РѕРїРµСЂР°С†РёРё (РєР°С‚РµРіРѕСЂРёСЏ РґРѕРєСѓРјРµРЅС‚Р°)",
            )
            idx_kaspi_payer_name = _pick_index(index, "payer_name", "payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_payer_iin = _pick_index(index, "iin_payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_payer_acc = _pick_index(index, "account_payer") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_name = _pick_index(index, "recipient_name", "recipient") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_iin = _pick_index(index, "iin_recipient") if parser_type == PARSER_KASPI else None
            idx_kaspi_recipient_acc = _pick_index(index, "account_recipient") if parser_type == PARSER_KASPI else None

            if idx_date is None or idx_currency is None or idx_debit is None or idx_credit is None:
                continue

            for row in rows[1:]:
                op_dt = _parse_operation_datetime(_get_cell(row, idx_date))
                if op_dt is None:
                    skipped += 1
                    continue

                debit = _to_float(_get_cell(row, idx_debit))
                credit = _to_float(_get_cell(row, idx_credit))
                amount_kzt = _to_float(_get_cell(row, idx_amount))
                amount_tenge = _safe_amount_tenge(amount_kzt, debit, credit)
                if amount_tenge <= 0:
                    skipped += 1
                    continue

                owner_name = _to_str(_get_cell(row, idx_owner_name), 255)
                owner_iin = _to_str(_get_cell(row, idx_owner_iin), 12)
                owner_acc = _to_str(_get_cell(row, idx_owner_acc), 20)
                if not owner_name or owner_name.upper() == "UNKNOWN":
                    owner_name = sheet_client_name or owner_name
                if (not owner_iin or owner_iin == "000000000000") and sheet_client_iin:
                    owner_iin = sheet_client_iin
                if not owner_acc and sheet_client_account:
                    owner_acc = sheet_client_account

                cp_name = _to_str(_get_cell(row, idx_cp_name), 255)
                cp_iin = _to_str(_get_cell(row, idx_cp_iin), 12)
                cp_acc = _to_str(_get_cell(row, idx_cp_acc), 20)
                direction_value = _to_str(_get_cell(row, idx_direction), 64)
                incoming = _is_incoming(direction_value, credit)

                kaspi_payer_name = _to_str(_get_cell(row, idx_kaspi_payer_name), 255) if idx_kaspi_payer_name is not None else ""
                kaspi_payer_iin = _to_str(_get_cell(row, idx_kaspi_payer_iin), 12) if idx_kaspi_payer_iin is not None else ""
                kaspi_payer_acc = _to_str(_get_cell(row, idx_kaspi_payer_acc), 20) if idx_kaspi_payer_acc is not None else ""
                kaspi_recipient_name = _to_str(_get_cell(row, idx_kaspi_recipient_name), 255) if idx_kaspi_recipient_name is not None else ""
                kaspi_recipient_iin = _to_str(_get_cell(row, idx_kaspi_recipient_iin), 12) if idx_kaspi_recipient_iin is not None else ""
                kaspi_recipient_acc = _to_str(_get_cell(row, idx_kaspi_recipient_acc), 20) if idx_kaspi_recipient_acc is not None else ""
                has_kaspi_explicit_parties = any(
                    [kaspi_payer_name, kaspi_payer_iin, kaspi_payer_acc, kaspi_recipient_name, kaspi_recipient_iin, kaspi_recipient_acc]
                )

                if parser_type == PARSER_KASPI and has_kaspi_explicit_parties:
                    sender_name = kaspi_payer_name
                    sender_iin = kaspi_payer_iin
                    sender_acc = kaspi_payer_acc
                    recipient_name = kaspi_recipient_name
                    recipient_iin = kaspi_recipient_iin
                    recipient_acc = kaspi_recipient_acc
                else:
                    sender_name = cp_name if incoming else owner_name
                    sender_iin = cp_iin if incoming else owner_iin
                    sender_acc = cp_acc if incoming else owner_acc
                    recipient_name = owner_name if incoming else cp_name
                    recipient_iin = owner_iin if incoming else cp_iin
                    recipient_acc = owner_acc if incoming else cp_acc

                if (not sender_name or sender_name.upper() == "UNKNOWN") and sheet_client_name:
                    sender_name = sheet_client_name
                if (not sender_iin or sender_iin == "000000000000") and sheet_client_iin:
                    sender_iin = sheet_client_iin
                if not sender_acc and sheet_client_account:
                    sender_acc = sheet_client_account

                out.append({
                    "date": op_dt,
                    "sender_name": sender_name or "UNKNOWN",
                    "sender_iin_bin": sender_iin or "000000000000",
                    "sender_account": sender_acc or "",
                    "recipient_name": recipient_name or "UNKNOWN",
                    "recipient_iin_bin": recipient_iin or "000000000000",
                    "recipient_account": recipient_acc or "",
                    "purpose": _to_str(_get_cell(row, idx_purpose), 1000),
                    "category": _fix_mojibake(_to_str(_get_cell(row, idx_category), 255)),
                    "operation_type": _fix_mojibake(_to_str(_get_cell(row, idx_operation_type), 255)) or _fix_mojibake(direction_value),
                    "currency": _to_str(_get_cell(row, idx_currency), 3).upper() or "KZT",
                    "debit": debit,
                    "credit": credit,
                    "amount_tenge": amount_tenge,
                })
        else:
            idx_date = _pick_index(index, "operation_date")
            idx_currency = _pick_index(index, "currency")
            idx_debit = _pick_index(index, "debit")
            idx_credit = _pick_index(index, "credit")
            if idx_date is None or idx_currency is None or idx_debit is None or idx_credit is None:
                continue

            idx_amount = _pick_index(index, "amount_kzt")
            idx_payer = _pick_index(index, "payer")
            idx_iin_payer = _pick_index(index, "iin_payer")
            idx_acc_payer = _pick_index(index, "account_payer")
            idx_recipient = _pick_index(index, "recipient")
            idx_iin_recipient = _pick_index(index, "iin_recipient")
            idx_acc_recipient = _pick_index(index, "account_recipient")
            idx_purpose = _pick_index(index, "transfer_purpose")
            idx_category = _pick_index(index, "category")
            idx_operation_type = _pick_index(
                index,
                "operation_type",
                "document_category",
                "op_type",
                "type",
                "РўРёРї РѕРїРµСЂР°С†РёРё",
            )

            for row in rows[1:]:
                op_dt = _parse_operation_datetime(_get_cell(row, idx_date))
                if op_dt is None:
                    skipped += 1
                    continue

                debit = _to_float(_get_cell(row, idx_debit))
                credit = _to_float(_get_cell(row, idx_credit))
                amount_kzt = _to_float(_get_cell(row, idx_amount))
                amount_tenge = _safe_amount_tenge(amount_kzt, debit, credit)
                if amount_tenge <= 0:
                    skipped += 1
                    continue

                out.append({
                    "date": op_dt,
                    "sender_name": _to_str(_get_cell(row, idx_payer), 255),
                    "sender_iin_bin": _to_str(_get_cell(row, idx_iin_payer), 12),
                    "sender_account": _to_str(_get_cell(row, idx_acc_payer), 20),
                    "recipient_name": _to_str(_get_cell(row, idx_recipient), 255),
                    "recipient_iin_bin": _to_str(_get_cell(row, idx_iin_recipient), 12),
                    "recipient_account": _to_str(_get_cell(row, idx_acc_recipient), 20),
                    "purpose": _to_str(_get_cell(row, idx_purpose), 1000),
                    "category": _fix_mojibake(_to_str(_get_cell(row, idx_category), 255)),
                    "operation_type": _fix_mojibake(_to_str(_get_cell(row, idx_operation_type), 255)),
                    "currency": _to_str(_get_cell(row, idx_currency), 3).upper() or "KZT",
                    "debit": debit,
                    "credit": credit,
                    "amount_tenge": amount_tenge,
                })

    return out, skipped


def _require_admin(authorization: Optional[str]) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing bearer token")

    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = decode_access_token(token)
    except Exception:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

    if payload.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    return payload


def _parser_type_to_source_bank(parser_type: str) -> Optional[str]:
    normalized = (parser_type or "").strip().lower()
    if normalized in {PARSER_KASPI, PARSER_KASPI_LEGACY}:
        return "kaspi"
    if normalized == PARSER_HALYK:
        return "halyk"
    if normalized in {PARSER_SMART, PARSER_BANK}:
        return None
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unknown parser_type. Use smart_parser, kaspi, kaspi_parser, bank_parser, halyk_parser or transactions_core",
    )


def _run_ingestion_pipeline(file_path: str, source_bank: Optional[str]):
    with IngestionPipeline() as pipeline:
        return pipeline.ingest_file(file_path, source_bank=source_bank)


async def _finalize_ingestion_import(file_id: Optional[str], uploader_email: str) -> int:
    if not file_id:
        return 0

    async with async_session() as session:
        uploaded_tx_ids = list(
            (
                await session.execute(
                    select(Transaction.id).where(Transaction.file_id == file_id)
                )
            ).scalars()
        )
        if not uploaded_tx_ids:
            return 0

        existing_meta = set(
            (
                await session.execute(
                    select(TransactionUploadMeta.tx_id).where(TransactionUploadMeta.tx_id.in_(uploaded_tx_ids))
                )
            ).scalars()
        )

        for tx_id in uploaded_tx_ids:
            if tx_id in existing_meta:
                continue
            session.add(
                TransactionUploadMeta(
                    tx_id=tx_id,
                    uploaded_by_email=uploader_email or "",
                    created_at=datetime.utcnow(),
                )
            )

        if len(existing_meta) != len(uploaded_tx_ids):
            await session.commit()

        return len(uploaded_tx_ids)


def _text_match_conditions(column, query: str):
    q = (query or "").strip()
    if not q:
        return None
    variants = {q, q.lower(), q.upper(), q.capitalize(), q.title()}
    like_conds = [column.ilike(f"%{v}%") for v in variants if v]
    return or_(*like_conds) if like_conds else None


def _non_empty_text_condition(column):
    return func.nullif(func.trim(func.coalesce(column, "")), "").isnot(None)


def _meaningful_transaction_condition():
    return or_(
        Transaction.date.isnot(None),
        Transaction.operation_date.isnot(None),
        func.coalesce(Transaction.amount_tenge, 0) > 0,
        func.coalesce(Transaction.debit, 0) > 0,
        func.coalesce(Transaction.credit, 0) > 0,
        _non_empty_text_condition(Transaction.sender_name),
        _non_empty_text_condition(Transaction.sender_account),
        _non_empty_text_condition(Transaction.recipient_name),
        _non_empty_text_condition(Transaction.recipient_account),
        _non_empty_text_condition(Transaction.purpose),
        _non_empty_text_condition(Transaction.currency),
    )


def _build_transaction_where_clause(
    date: Optional[str],
    category: Optional[str],
    search: Optional[str],
    min_amount: Optional[float],
    max_amount: Optional[float],
    currency: Optional[str],
    sender: Optional[str],
    recipient: Optional[str],
):
    conditions = [_meaningful_transaction_condition()]
    if date:
        dt = _parse_date(date)
        day_start = dt.replace(hour=0, minute=0, second=0)
        day_end = dt.replace(hour=23, minute=59, second=59)
        conditions.append(Transaction.date.between(day_start, day_end))

    if category:
        category_q = category.strip()
        category_mojibake = _to_mojibake(category_q)
        category_expr = _derived_category_expr()
        cond_main = _text_match_conditions(category_expr, category_q)
        if cond_main is not None:
            if category_mojibake and category_mojibake != category_q:
                cond_moji = _text_match_conditions(category_expr, category_mojibake)
                if cond_moji is not None:
                    conditions.append(or_(cond_main, cond_moji))
                else:
                    conditions.append(cond_main)
            else:
                conditions.append(cond_main)

    if search:
        search_cond = _text_match_conditions(Transaction.purpose, search)
        if search_cond is not None:
            conditions.append(search_cond)
    if min_amount is not None:
        conditions.append(Transaction.amount_tenge >= min_amount)
    if max_amount is not None:
        conditions.append(Transaction.amount_tenge <= max_amount)
    if currency:
        conditions.append(Transaction.currency == currency.upper())
    if sender:
        sender_q = sender.strip()
        sender_name_cond = _text_match_conditions(Transaction.sender_name, sender_q)
        sender_match_list = [
            Transaction.sender_iin_bin.ilike(f"%{sender_q}%"),
            Transaction.sender_account.ilike(f"%{sender_q}%"),
        ]
        if sender_name_cond is not None:
            sender_match_list.insert(0, sender_name_cond)
        conditions.append(
            or_(*sender_match_list)
        )
    if recipient:
        recipient_q = recipient.strip()
        recipient_name_cond = _text_match_conditions(Transaction.recipient_name, recipient_q)
        recipient_match_list = [
            Transaction.recipient_iin_bin.ilike(f"%{recipient_q}%"),
            Transaction.recipient_account.ilike(f"%{recipient_q}%"),
        ]
        if recipient_name_cond is not None:
            recipient_match_list.insert(0, recipient_name_cond)
        conditions.append(
            or_(*recipient_match_list)
        )
    return and_(*conditions) if conditions else True


def _build_transaction_order_by(sort_by: str, sort_dir: str):
    key = (sort_by or "date").strip()
    direction = (sort_dir or "desc").strip().lower()
    is_desc = direction != "asc"

    text_exprs = {
        "category": func.lower(func.coalesce(_derived_category_expr(), "")),
        "operationType": func.lower(func.coalesce(Transaction.operation_type, "")),
        "purpose": func.lower(func.coalesce(Transaction.purpose, "")),
        "uploadedBy": func.lower(func.coalesce(TransactionUploadMeta.uploaded_by_email, "")),
        "sender": func.lower(
            func.coalesce(
                func.nullif(func.trim(Transaction.sender_name), ""),
                func.nullif(func.trim(Transaction.sender_account), ""),
                Transaction.sender_iin_bin,
                "",
            )
        ),
        "recipient": func.lower(
            func.coalesce(
                func.nullif(func.trim(Transaction.recipient_name), ""),
                func.nullif(func.trim(Transaction.recipient_account), ""),
                Transaction.recipient_iin_bin,
                "",
            )
        ),
    }

    numeric_exprs = {
        "date": Transaction.date,
        "debit": Transaction.debit,
        "credit": Transaction.credit,
        "amountTenge": Transaction.amount_tenge,
    }

    if key == "currency":
        # asc: KZT first, desc: other currencies first
        rank_expr = case((Transaction.currency == "KZT", 0), else_=1)
        if is_desc:
            return [rank_expr.desc(), Transaction.currency.asc(), Transaction.date.desc()]
        return [rank_expr.asc(), Transaction.currency.asc(), Transaction.date.desc()]

    expr = text_exprs.get(key)
    if expr is None:
        expr = numeric_exprs.get(key)
    if expr is None:
        expr = Transaction.date
    primary = expr.desc() if is_desc else expr.asc()
    return [primary, Transaction.date.desc()]


@router.get("", response_model=TransactionListResponse)
async def list_transactions(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    category: Optional[str] = Query(None, description="Transaction category"),
    search: Optional[str] = Query(None, description="Full-text search on purpose"),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Sender IIN/BIN/Account"),
    recipient: Optional[str] = Query(None, description="Recipient IIN/BIN/Account"),
    sort_by: str = Query("date"),
    sort_dir: str = Query("desc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    order_by = _build_transaction_order_by(sort_by, sort_dir)
    where_clause = _build_transaction_where_clause(
        date=date,
        category=category,
        search=search,
        min_amount=min_amount,
        max_amount=max_amount,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )

    count_q = select(func.count(Transaction.id)).where(where_clause)
    total = (await db.execute(count_q)).scalar() or 0

    sum_q = select(
        func.coalesce(func.sum(Transaction.debit), 0),
        func.coalesce(func.sum(Transaction.credit), 0),
    ).where(where_clause)
    sums = (await db.execute(sum_q)).one()
    total_debit, total_credit = float(sums[0]), float(sums[1])

    offset = (page - 1) * per_page
    rows_q = (
        select(Transaction, TransactionUploadMeta.uploaded_by_email.label("uploaded_by_email"))
        .outerjoin(TransactionUploadMeta, TransactionUploadMeta.tx_id == Transaction.id)
        .where(where_clause)
        .order_by(*order_by)
        .offset(offset)
        .limit(per_page)
    )
    rows = (await db.execute(rows_q)).all()
    total_pages = max(1, (total + per_page - 1) // per_page)

    data = [
        TransactionOut(
            id=str(t.id),
            date=_format_transaction_dt(t),
            sender=CounterpartyOut(
                name=_fix_mojibake(t.sender_name) or "",
                iin_bin=t.sender_iin_bin or "",
                account=t.sender_account or "",
            ),
            recipient=CounterpartyOut(
                name=_fix_mojibake(t.recipient_name) or "",
                iin_bin=t.recipient_iin_bin or "",
                account=t.recipient_account or "",
            ),
            category=_derive_display_category(t.category or "", t.purpose or "", t.operation_type or "", t.direction or ""),
            operation_type=_fix_mojibake(t.operation_type) or "",
            purpose=_fix_mojibake(t.purpose) or "",
            currency=t.currency or "",
            debit=float(t.debit or 0),
            credit=float(t.credit or 0),
            amount_tenge=float(t.amount_tenge or 0),
            uploaded_by_email=uploaded_by_email or "",
        )
        for t, uploaded_by_email in rows
    ]

    return TransactionListResponse(
        data=data,
        pagination=PaginationOut(page=page, per_page=per_page, total=total, total_pages=total_pages),
        summary=SummaryOut(total_debit=total_debit, total_credit=total_credit),
    )


@router.get("/export")
async def export_transactions(
    date: Optional[str] = Query(None, description="Exact date DD.MM.YYYY"),
    category: Optional[str] = Query(None, description="Transaction category"),
    search: Optional[str] = Query(None, description="Full-text search on purpose"),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    currency: Optional[str] = Query(None),
    sender: Optional[str] = Query(None, description="Sender IIN/BIN/Account"),
    recipient: Optional[str] = Query(None, description="Recipient IIN/BIN/Account"),
    db: AsyncSession = Depends(get_db),
):
    where_clause = _build_transaction_where_clause(
        date=date,
        category=category,
        search=search,
        min_amount=min_amount,
        max_amount=max_amount,
        currency=currency,
        sender=sender,
        recipient=recipient,
    )

    rows_q = (
        select(Transaction, TransactionUploadMeta.uploaded_by_email.label("uploaded_by_email"))
        .outerjoin(TransactionUploadMeta, TransactionUploadMeta.tx_id == Transaction.id)
        .where(where_clause)
        .order_by(Transaction.date.desc())
    )
    rows = (await db.execute(rows_q)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append([
        "Категория",
        "Вид операции",
        "Дата",
        "Отправитель",
        "ИИН/БИН отправителя",
        "Счет отправителя",
        "Получатель",
        "ИИН/БИН получателя",
        "Счет получателя",
        "Назначение",
        "Валюта",
        "Расход",
        "Поступление",
        "Сумма (тенге)",
        "Кто добавил",
    ])

    for t, uploaded_by_email in rows:
        ws.append([
            _derive_display_category(t.category or "", t.purpose or "", t.operation_type or "", t.direction or ""),
            _fix_mojibake(t.operation_type or ""),
            _format_transaction_dt(t),
            _fix_mojibake(t.sender_name or ""),
            t.sender_iin_bin or "",
            t.sender_account or "",
            _fix_mojibake(t.recipient_name or ""),
            t.recipient_iin_bin or "",
            t.recipient_account or "",
            _fix_mojibake(t.purpose or ""),
            t.currency or "",
            float(t.debit or 0),
            float(t.credit or 0),
            float(t.amount_tenge or 0),
            uploaded_by_email or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    now_label = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transactions_{now_label}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import-statement", response_model=TransactionImportResponse)
async def import_statement(
    file: UploadFile = File(...),
    parser_type: str = Form(PARSER_KASPI),
    authorization: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    auth_payload = _require_admin(authorization)
    uploader_email = str(auth_payload.get("email") or "").strip().lower()
    parser_type = (parser_type or PARSER_KASPI).strip().lower()

    filename = (file.filename or "").lower()
    allowed_ext = (".xls", ".xlsx")
    if parser_type == PARSER_TRANSACTIONS_CORE:
        allowed_ext = (".csv",)

    if not filename.endswith(allowed_ext):
        if parser_type == PARSER_TRANSACTIONS_CORE:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .csv for transactions_core")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload .xls or .xlsx")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")

    if parser_type == PARSER_TRANSACTIONS_CORE:
        tx_rows, skipped = _extract_transactions_from_transactions_core_csv(content)
        inserted = 0
        file_uuid = str(uuid4())
        statement_uuid = str(uuid4())
        format_uuid = str(uuid4())
        source_sheet = file.filename or "transactions_core.csv"
        for idx, tx in enumerate(tx_rows, start=2):
            payload = _build_core_transaction_payload(
                tx,
                uploader_email=uploader_email,
                source_bank="transactions_core",
                source_sheet=source_sheet,
                source_row_no=idx,
                file_id=file_uuid,
                statement_id=statement_uuid,
                format_id=format_uuid,
                raw_row_json=tx,
            )
            db.add(Transaction(**payload))
            db.add(TransactionUploadMeta(
                tx_id=payload["id"],
                uploaded_by_email=uploader_email or "",
                created_at=datetime.utcnow(),
            ))
            inserted += 1

        await db.commit()
        return TransactionImportResponse(
            inserted=inserted,
            skipped=skipped,
            message=f"Imported {inserted} transactions from transactions_core CSV",
        )
    source_bank = _parser_type_to_source_bank(parser_type)

    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # The sync ingestion pipeline needs its own psycopg2 connection. Release
        # the request-scoped async connection first so tiny Postgres instances
        # don't fail with "remaining connection slots are reserved".
        maybe = db.close()
        if inspect.isawaitable(maybe):
            await maybe
        maybe = async_engine.dispose()
        if inspect.isawaitable(maybe):
            await maybe
        result = await run_in_threadpool(_run_ingestion_pipeline, tmp_path, source_bank)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Ingestion pipeline failed: {exc}") from exc
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    parsed_rows = int(result.get("core_rows") or 0)
    inserted = await _finalize_ingestion_import(result.get("file_id"), uploader_email)
    skipped = max(parsed_rows - inserted, 0)

    bank_label = result.get("bank") or source_bank or "auto"
    return TransactionImportResponse(
        inserted=inserted,
        skipped=skipped,
        message=f"Imported {inserted} transactions via {bank_label}",
    )


