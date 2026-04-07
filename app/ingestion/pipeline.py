from __future__ import annotations

import logging
import os
import uuid
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from app.config import settings
from app.database import engine
from app.db.schema import ensure_schema
from app.db.writers import PostgresWriter

from app.ingestion.extractor.universal_extractor import ExcelUniversalExtractor
from app.ingestion.extractor.dataframe_cleaner import clean_dataframe
from app.ingestion.metadata.statement_meta_extractor import StatementMetadataExtractor
from app.ingestion.mapping.canonical_mapper import CanonicalMapper
from app.ingestion.mapping.embedding_mapper import EmbeddingBackend
from app.ingestion.registry.format_registry import FormatRegistryService
from app.ingestion.registry.discovery_logger import DiscoveryLogger

from app.ingestion.extractor.adapter_loader import load_adapters
from app.classification import CategoryService
from app.utils.hashing import sha256_file

log = logging.getLogger(__name__)

AdapterDF = Union[pd.DataFrame, Tuple[pd.DataFrame, Dict[str, Any]]]

# ── embedding batch size ──────────────────────────────────────────────────────
# Adjust down if GPU memory is constrained.
_EMBED_BATCH_SIZE = 256


def _vec_to_pg_literal(vec: np.ndarray) -> str:
    """Convert a float32 numpy vector to a pgvector literal string."""
    arr = np.asarray(vec, dtype=np.float32).reshape(-1)
    return "[" + ",".join(f"{v:.6f}" for v in arr) + "]"


def _attach_embeddings(rows: List[Dict[str, Any]], embedder: EmbeddingBackend) -> None:
    """
    Batch-embed semantic_text for all rows in-place.

    Rows that have no semantic_text get semantic_embedding = None.
    Skips silently if embedder is disabled.
    """
    if not embedder.enabled:
        return

    # collect indices that have text to embed
    idx_to_embed = [
        i for i, r in enumerate(rows)
        if r.get("semantic_text")
    ]
    if not idx_to_embed:
        return

    texts = [rows[i]["semantic_text"] for i in idx_to_embed]

    # process in batches to control memory
    for batch_start in range(0, len(texts), _EMBED_BATCH_SIZE):
        batch_texts = texts[batch_start: batch_start + _EMBED_BATCH_SIZE]
        batch_idx   = idx_to_embed[batch_start: batch_start + _EMBED_BATCH_SIZE]

        try:
            vecs = embedder.embed(batch_texts)   # shape: (N, dim)
            for list_idx, vec in zip(batch_idx, vecs):
                rows[list_idx]["semantic_embedding"] = _vec_to_pg_literal(vec)
        except Exception:
            log.exception(
                "Embedding batch %d-%d failed — semantic_embedding will be NULL",
                batch_start, batch_start + len(batch_texts),
            )


class IngestionPipeline:
    def __init__(self):
        self.settings = settings
        self.adapters = load_adapters()

        self.engine = engine
        ensure_schema(self.engine)
        self.writer = PostgresWriter(self.engine, parser_version=self.settings.PARSER_VERSION)

        self.embedder = EmbeddingBackend(
            settings.EMBEDDING_MODEL_PATH,
            provider=settings.AFM_EMBEDDING_PROVIDER,
            ollama_base_url=settings.AFM_EMBEDDING_BASE_URL,
            ollama_timeout_s=settings.AFM_EMBEDDING_TIMEOUT_S,
        )
        self.mapper = CanonicalMapper(self.embedder, threshold=self.settings.EMBEDDING_THRESHOLD)
        self.category_service = CategoryService(self.embedder)

        self.format_registry = FormatRegistryService(
            writer=self.writer,
            embedder=self.embedder,
            similarity_threshold=settings.FORMAT_SIMILARITY_THRESHOLD,
        )
        self.discovery = DiscoveryLogger(self.writer)

        self.extractor = ExcelUniversalExtractor()
        self.meta_extractor = StatementMetadataExtractor()

        log.info(
            "IngestionPipeline ready  embedder=%s  semantic_embedding=%s",
            "ON" if self.embedder.enabled else "OFF",
            "will be computed" if self.embedder.enabled else "NULL (no model)",
        )

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            self.engine.dispose()
        except Exception:
            log.exception("Failed to dispose sync ingestion engine")
        return False

    def _find_adapter(self, bank: str):
        for a in self.adapters:
            if getattr(a, "bank_name", None) == bank:
                return a
        return None

    # ── folder ingestion ──────────────────────────────────────────────────────

    def ingest_data_folder(self, data_root: str) -> List[Dict[str, Any]]:
        log.info("Scanning data folder: %s", data_root)
        if not os.path.isdir(data_root):
            raise FileNotFoundError(f"data_root not found: {data_root}")

        results: List[Dict[str, Any]] = []
        for adapter in self.adapters:
            bank = adapter.bank_name
            files = adapter.list_files(data_root)
            if not files:
                log.info("No files for bank=%s in %s", bank, os.path.join(data_root, bank))
                continue
            log.info("bank=%s -> %d file(s)", bank, len(files))
            for f in sorted(files):
                try:
                    results.append(self.ingest_file(f, source_bank=bank))
                except Exception:
                    log.exception("Failed ingesting file=%s (bank=%s)", f, bank)
        return results

    # ── single file ingestion ─────────────────────────────────────────────────

    def ingest_file(
        self,
        xlsx_path: str,
        source_bank: Optional[str] = None,
        project_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        file_id  = str(uuid.uuid4())
        filename = os.path.basename(xlsx_path)
        bank     = source_bank or "unknown"
        checksum = sha256_file(xlsx_path)

        self.writer.insert_raw_file(
            file_id=file_id,
            project_id=project_id,
            source_bank=bank,
            filename=filename,
            sha256=checksum,
        )

        all_core: List[Dict[str, Any]] = []
        all_ext:  List[Dict[str, Any]] = []
        all_disc: List[Dict[str, Any]] = []
        statements_count = 0

        adapter = self._find_adapter(bank)
        extracted: List[AdapterDF] = []
        if adapter is None and bank == "unknown":
            for candidate in self.adapters:
                try:
                    candidate_rows: List[AdapterDF] = candidate.extract(xlsx_path) or []
                except Exception:
                    log.exception("Adapter auto-detect failed for bank=%s file=%s", candidate.bank_name, filename)
                    continue
                if candidate_rows:
                    adapter = candidate
                    extracted = candidate_rows
                    bank = candidate.bank_name
                    log.info("Smart parser auto-detected adapter=%s for file=%s", bank, filename)
                    break

        # ── CASE 1: bank-specific adapter ─────────────────────────────────────
        used_adapter = False
        if adapter is not None:
            log.info("Using adapter=%s for file=%s", adapter.bank_name, filename)
            if not extracted:
                extracted = adapter.extract(xlsx_path) or []
            if extracted:
                used_adapter = True
            else:
                log.warning(
                    "Adapter %s returned 0 tables for %s -> fallback to universal",
                    adapter.bank_name, filename,
                )

            for idx, item in enumerate(extracted, start=1):
                if isinstance(item, tuple) and len(item) == 2:
                    df, stmt_meta_from_adapter = item
                    stmt_meta_from_adapter = stmt_meta_from_adapter or {}
                else:
                    df = item  # type: ignore[assignment]
                    stmt_meta_from_adapter = {}

                if not isinstance(df, pd.DataFrame):
                    log.warning(
                        "Adapter %s returned non-DF item #%d for %s. Skipping.",
                        bank, idx, filename,
                    )
                    continue

                df = clean_dataframe(df)
                if df.empty or len(df.columns) < 3:
                    continue

                format_id = self.format_registry.register_or_get_format(
                    source_bank=bank,
                    headers=list(df.columns),
                )

                statement_id = str(uuid.uuid4())
                stmt_row = {
                    "statement_id":   statement_id,
                    "file_id":        file_id,
                    "project_id":     project_id,
                    "source_bank":    bank,
                    "source_sheet":   stmt_meta_from_adapter.get("source_sheet"),
                    "source_block_id": stmt_meta_from_adapter.get("source_block_id", idx),
                    "format_id":      format_id,
                    "client_name":    stmt_meta_from_adapter.get("client_name"),
                    "client_iin_bin": stmt_meta_from_adapter.get("client_iin_bin"),
                    "contract_no":    stmt_meta_from_adapter.get("contract_no"),
                    "account_iban":   stmt_meta_from_adapter.get("account_iban"),
                    "account_type":   stmt_meta_from_adapter.get("account_type"),
                    "currency":       stmt_meta_from_adapter.get("currency"),
                    "statement_date": stmt_meta_from_adapter.get("statement_date"),
                    "period_from":    stmt_meta_from_adapter.get("period_from"),
                    "period_to":      stmt_meta_from_adapter.get("period_to"),
                    "opening_balance": stmt_meta_from_adapter.get("opening_balance"),
                    "closing_balance": stmt_meta_from_adapter.get("closing_balance"),
                    "total_debit":    stmt_meta_from_adapter.get("total_debit"),
                    "total_credit":   stmt_meta_from_adapter.get("total_credit"),
                    "meta_json":      stmt_meta_from_adapter.get("meta_json") or {"source": "adapter"},
                }
                self.writer.insert_statement(stmt_row)
                statements_count += 1

                mapped, unmapped = self.mapper.map_headers(df)
                ctx = {
                    "file_id":          file_id,
                    "statement_id":     statement_id,
                    "format_id":        format_id,
                    "project_id":       project_id,
                    "source_bank":      bank,
                    "source_sheet":     stmt_meta_from_adapter.get("source_sheet"),
                    "source_block_id":  stmt_meta_from_adapter.get("source_block_id", idx),
                    "source_row_base":  int(stmt_meta_from_adapter.get("source_row_base", 0)),
                    "store_raw_row_json": self.settings.store_raw_row_json,
                    "account_iban":     stmt_meta_from_adapter.get("account_iban"),
                }

                core_rows, ext_rows, disc_rows = self.mapper.to_rows(df, mapped, ctx)
                all_core.extend(core_rows)
                all_ext.extend(ext_rows)
                if unmapped:
                    all_disc.extend(disc_rows)

        # ── CASE 2: universal fallback ─────────────────────────────────────────
        if (adapter is None) or (not used_adapter):
            log.info("Universal extractor for file=%s (bank=%s)", filename, bank)
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)

            for sheet_name in wb.sheetnames:
                grid = self.extractor.load_sheet_grid(xlsx_path, sheet_name)
                blocks = self.extractor.detect_blocks(grid, sheet_name)
                if not blocks:
                    continue

                for bidx, block in enumerate(blocks, start=1):
                    df = self.extractor.extract_block_df(grid, block)
                    df = clean_dataframe(df)
                    if df.empty or len(df.columns) < 3:
                        continue

                    format_id = self.format_registry.register_or_get_format(
                        source_bank=bank,
                        headers=list(df.columns),
                    )

                    stmt_meta = self.meta_extractor.extract_for_block(
                        grid=grid,
                        block=block,
                        source_bank=bank,
                        max_lookback_rows=self.settings.max_meta_lookback_rows,
                    )

                    statement_id = str(uuid.uuid4())
                    self.writer.insert_statement({
                        "statement_id":   statement_id,
                        "file_id":        file_id,
                        "project_id":     project_id,
                        "source_bank":    bank,
                        "source_sheet":   sheet_name,
                        "source_block_id": bidx,
                        "format_id":      format_id,
                        **stmt_meta,
                    })
                    statements_count += 1

                    mapped, unmapped = self.mapper.map_headers(df)
                    ctx = {
                        "file_id":          file_id,
                        "statement_id":     statement_id,
                        "format_id":        format_id,
                        "project_id":       project_id,
                        "source_bank":      bank,
                        "source_sheet":     sheet_name,
                        "source_block_id":  bidx,
                        "source_row_base":  block.data_start_row_idx,
                        "store_raw_row_json": self.settings.store_raw_row_json,
                        "account_iban":     stmt_meta.get("account_iban"),
                    }

                    core_rows, ext_rows, disc_rows = self.mapper.to_rows(df, mapped, ctx)
                    all_core.extend(core_rows)
                    all_ext.extend(ext_rows)
                    if unmapped:
                        all_disc.extend(disc_rows)

        # ── SEMANTIC EMBEDDING (batch, before DB insert) ───────────────────────
        if all_core:
            for row in all_core:
                row["project_id"] = project_id
            try:
                self.category_service.classify_rows(all_core)
            except Exception:
                log.exception("Category classification failed; continuing without blocking upload")
            log.info(
                "Computing semantic_embedding for %d rows (embedder=%s)…",
                len(all_core),
                "ON" if self.embedder.enabled else "OFF",
            )
            _attach_embeddings(all_core, self.embedder)

        # ── DB inserts ─────────────────────────────────────────────────────────
        self.writer.bulk_insert_core_dedup(all_core)
        self.writer.bulk_insert_ext(all_ext)
        if all_disc:
            self.discovery.log(all_disc)
        self.writer.mark_parsed(file_id=file_id)

        result = {
            "file_id":        file_id,
            "bank":           bank,
            "filename":       filename,
            "statements":     statements_count,
            "core_rows":      len(all_core),
            "ext_rows":       len(all_ext),
            "discovery_cols": len({d["raw_column_name"] for d in all_disc}) if all_disc else 0,
            "semantic_embedded": self.embedder.enabled,
        }
        log.info("Ingested %s -> %s", filename, result)
        return result
