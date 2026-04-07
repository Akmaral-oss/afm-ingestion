"""
app/ingestion/extractor/universal_extractor.py

Supports both .xlsx (openpyxl) and legacy .xls (xlrd).
.xls files are converted to a grid in-memory via pandas — no temp files needed.
"""
from __future__ import annotations

from typing import Any, List
from app.ingestion.extractor.block_detector import detect_blocks, DetectedBlock, build_df_from_block


def _is_xls(path: str) -> bool:
    return path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")


def _load_xlsx_grid(path: str, sheet_name: str) -> List[List[Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    max_col = ws.max_column
    return [list(row[:max_col]) for row in ws.iter_rows(values_only=True)]


def _load_xls_grid(path: str, sheet_name: str) -> List[List[Any]]:
    import xlrd  # type: ignore
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name)
    grid = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
            if cell.ctype == 0:
                row.append(None)
            elif cell.ctype == 3:
                # Convert xlrd date float to Python date
                import xlrd as _xlrd
                import datetime
                dt = _xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(dt)
            else:
                row.append(cell.value)
        grid.append(row)
    return grid


def _get_sheet_names(path: str) -> List[str]:
    if _is_xls(path):
        import xlrd  # type: ignore
        return xlrd.open_workbook(path).sheet_names()
    else:
        from openpyxl import load_workbook
        return load_workbook(path, read_only=True, data_only=True).sheetnames


class ExcelUniversalExtractor:
    def load_sheet_grid(self, path: str, sheet_name: str) -> List[List[Any]]:
        if _is_xls(path):
            return _load_xls_grid(path, sheet_name)
        return _load_xlsx_grid(path, sheet_name)

    def get_sheet_names(self, path: str) -> List[str]:
        return _get_sheet_names(path)

    def detect_blocks(self, grid: List[List[Any]], sheet_name: str) -> List[DetectedBlock]:
        return detect_blocks(grid, sheet_name)

    def extract_block_df(self, grid: List[List[Any]], block: DetectedBlock):
        return build_df_from_block(grid, block)
