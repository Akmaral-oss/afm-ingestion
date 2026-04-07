from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

import os
import pandas as pd
from openpyxl import load_workbook

from app.ingestion.adapters.base_adapter import BankAdapter
from app.ingestion.extractor.block_detector import DetectedBlock
from app.ingestion.metadata.statement_meta_extractor import StatementMetadataExtractor
from app.utils.text_utils import norm_text


@dataclass
class SimpleBlock:
    sheet_name: str
    header_row_idx: int
    data_start_row_idx: int
    data_end_row_idx: int


HALYK_HEADER_MARKERS = [
    "дата и время операции",
    "дата операции",
    "дата/время",
    "валюта операции",
    "сумма в тенге",
    # Some files still surface mojibake instead of proper Cyrillic.
    "РґР°С‚Р° Рё РІСЂРµРјСЏ РѕРїРµСЂР°С†РёРё",
    "РґР°С‚Р° РѕРїРµСЂР°С†РёРё",
    "РґР°С‚Р°/РІСЂРµРјСЏ",
    "РІР°Р»СЋС‚Р° РѕРїРµСЂР°С†РёРё",
    "СЃСѓРјРјР° РІ С‚РµРЅРіРµ",
]


def _is_xls(path: str) -> bool:
    return path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")


def _get_sheet_names(path: str) -> List[str]:
    if _is_xls(path):
        import xlrd

        wb = xlrd.open_workbook(path)
        return wb.sheet_names()

    wb = load_workbook(path, data_only=True, read_only=True)
    return wb.sheetnames


def _load_xlsx_grid(path: str, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    grid: List[List[Any]] = []
    max_col = ws.max_column
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row[:max_col]))
    return grid


def _load_xls_grid(path: str, sheet_name: str) -> List[List[Any]]:
    import xlrd

    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name)
    grid: List[List[Any]] = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            row.append(cell.value if cell.ctype not in (0, 5, 6) else None)
        grid.append(row)
    return grid


class HalykAdapter(BankAdapter):
    bank_name = "halyk"

    def __init__(self):
        self.meta = StatementMetadataExtractor()

    def list_files(self, data_root: str) -> List[str]:
        path = os.path.join(data_root, "halyk")
        if not os.path.exists(path):
            return []
        return [
            os.path.join(path, f)
            for f in os.listdir(path)
            if f.lower().endswith(".xlsx") or f.lower().endswith(".xls")
        ]

    def load_grid(self, path: str, sheet_name: str) -> List[List[Any]]:
        if _is_xls(path):
            return _load_xls_grid(path, sheet_name)
        return _load_xlsx_grid(path, sheet_name)

    def _row_join(self, row: List[Any]) -> str:
        return " ".join(norm_text(x) for x in row if x is not None)

    def _is_header_row(self, row: List[Any]) -> bool:
        joined = self._row_join(row)
        if not joined:
            return False
        return any(marker in joined for marker in HALYK_HEADER_MARKERS)

    def _is_empty_row(self, row: List[Any]) -> bool:
        return all(norm_text(x) == "" for x in row)

    def _scan_end(self, grid: List[List[Any]], start_idx: int) -> int:
        n = len(grid)
        empty_streak = 0
        last_data = start_idx - 1

        for r in range(start_idx, n):
            row = grid[r]

            if self._is_header_row(row):
                break

            if self._is_empty_row(row):
                empty_streak += 1
            else:
                empty_streak = 0
                last_data = r

            if empty_streak >= 3:
                break

        return max(last_data, start_idx - 1)

    def _detect_blocks(self, grid: List[List[Any]], sheet_name: str) -> List[SimpleBlock]:
        blocks: List[SimpleBlock] = []
        i = 0
        n = len(grid)

        while i < n:
            if self._is_header_row(grid[i]):
                header_i = i
                data_start = i + 1
                data_end = self._scan_end(grid, data_start)

                if data_end >= data_start:
                    blocks.append(
                        SimpleBlock(
                            sheet_name=sheet_name,
                            header_row_idx=header_i,
                            data_start_row_idx=data_start,
                            data_end_row_idx=data_end,
                        )
                    )
                    i = data_end + 1
                    continue
            i += 1

        return blocks

    def _build_df(self, grid: List[List[Any]], block: SimpleBlock) -> pd.DataFrame:
        header = [norm_text(x) for x in grid[block.header_row_idx]]
        data = grid[block.data_start_row_idx:block.data_end_row_idx + 1]

        df = pd.DataFrame(data, columns=header)
        df = df.dropna(axis=0, how="all")
        df = df.dropna(axis=1, how="all")

        return df

    def extract(self, file_path: str) -> List[Tuple[pd.DataFrame, Dict[str, Any]]]:
        out: List[Tuple[pd.DataFrame, Dict[str, Any]]] = []

        for sheet_name in _get_sheet_names(file_path):
            grid = self.load_grid(file_path, sheet_name)
            blocks = self._detect_blocks(grid, sheet_name)

            for bidx, b in enumerate(blocks, start=1):
                df = self._build_df(grid, b)

                if df.empty or len(df.columns) < 2:
                    continue

                block = DetectedBlock(
                    sheet_name=sheet_name,
                    header_row_idx=b.header_row_idx,
                    data_start_row_idx=b.data_start_row_idx,
                    data_end_row_idx=b.data_end_row_idx,
                    header_rows=1,
                    group_row_idx=None,
                    subheader_row_idx=None,
                )

                stmt_meta = self.meta.extract_for_block(
                    grid=grid,
                    block=block,
                    source_bank=self.bank_name,
                    max_lookback_rows=30,
                    max_lookahead_rows=40,
                    tail_rows_in_block=40,
                )

                stmt_meta["source_sheet"] = sheet_name
                stmt_meta["source_block_id"] = bidx
                stmt_meta["source_row_base"] = b.data_start_row_idx
                stmt_meta["meta_json"] = {
                    **(stmt_meta.get("meta_json") or {}),
                    "source": "halyk_adapter",
                    "header_row_idx": b.header_row_idx,
                    "data_start": b.data_start_row_idx,
                    "data_end": b.data_end_row_idx,
                }

                out.append((df, stmt_meta))

        return out
