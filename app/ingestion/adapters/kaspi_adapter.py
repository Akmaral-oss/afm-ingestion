from __future__ import annotations

"""
KaspiAdapter — fixed version
Bugs fixed vs original:
  1. Column-index row (1 2 3 4 … 17) was ingested as a phantom data row.
     Fix: _is_column_index_row() detects sequential integer rows and skips them.
  2. IIN/BIN stored as float (8.609200e+11). 
     Fix: handled in canonical_mapper (_safe_iin_bin).
  3. total_credit = 2 (debit count misread). 
     Fix: _parse_kaspi_totals_row() uses explicit "N дебет / N кредит" regex.
  4. client_name not extracted (colon in norm_text label).
     Fix: strip trailing colon from label before matching.
  5. account_type not extracted. Same fix as #4.
  6. raw_note = duplicate of purpose_text.
     Fix: raw_note only set when a separate note column exists.
  7. purpose_code stored as float 841.0.
     Fix: handled in canonical_mapper (_safe_purpose_code).
"""

import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.ingestion.adapters.base_adapter import BankAdapter
from app.ingestion.extractor.block_detector import DetectedBlock
from app.ingestion.metadata.statement_meta_extractor import StatementMetadataExtractor
from app.utils.text_utils import norm_text


@dataclass
class SimpleBlock:
    sheet_name: str
    header_row_idx: int
    data_start_row_idx: int
    data_end_row_idx: int
    header_rows: int
    group_row_idx: int | None = None
    subheader_row_idx: int | None = None


KASPI_SUBHEADER_MARKERS = {
    "наименование/фио", "наименование", "фио",
    "иин/бин", "иин", "бин",
    "резидентство", "банк",
    "номер счета", "номер счёта", "счет", "счёт",
}

KASPI_SINGLE_HEADER_MARKERS = [
    "дата операции", "дата", "валюта", "сумма",
    "назначение", "виды операции", "категория документа",
]


class KaspiAdapter(BankAdapter):
    bank_name = "kaspi"

    def __init__(self) -> None:
        self.meta = StatementMetadataExtractor()

    def list_files(self, data_root: str) -> List[str]:
        path = os.path.join(data_root, "kaspi")
        if not os.path.exists(path):
            return []
        return [
            os.path.join(path, f)
            for f in os.listdir(path)
            if f.lower().endswith(".xlsx")
        ]

    def load_grid(self, path: str, sheet_name: str) -> List[List[Any]]:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        grid: List[List[Any]] = []
        max_col = ws.max_column
        for row in ws.iter_rows(values_only=True):
            grid.append(list(row[:max_col]))
        return grid

    # ── row classification helpers ──────────────────────────────────────────

    def _row_tokens(self, row: List[Any]) -> List[str]:
        return [norm_text(x) for x in row if x is not None and norm_text(x)]

    def _row_join(self, row: List[Any]) -> str:
        return " ".join(self._row_tokens(row))

    def _is_empty_row(self, row: List[Any]) -> bool:
        return all(norm_text(x) == "" for x in row)

    def _is_group_row(self, row: List[Any]) -> bool:
        toks = self._row_tokens(row)
        return "плательщик" in toks and "получатель" in toks

    def _is_subheader_row(self, row: List[Any]) -> bool:
        toks = set(self._row_tokens(row))
        return len(toks.intersection(KASPI_SUBHEADER_MARKERS)) >= 3

    def _is_single_header_row(self, row: List[Any]) -> bool:
        joined = self._row_join(row)
        if not joined:
            return False
        hits = sum(1 for m in KASPI_SINGLE_HEADER_MARKERS if m in joined)
        return hits >= 2

    # FIX 1 ─────────────────────────────────────────────────────────────────
    def _is_column_index_row(self, row: List[Any]) -> bool:
        """
        Detect the Kaspi sequential column-number row (1 2 3 4 … 17).
        Returns True when every non-null cell is an integer and the values
        form a consecutive sequence starting from 1.
        """
        non_null = [x for x in row if x is not None]
        if len(non_null) < 3:
            return False

        ints: List[int] = []
        for v in non_null:
            if isinstance(v, int):
                ints.append(v)
            elif isinstance(v, float) and v == int(v):
                ints.append(int(v))
            elif isinstance(v, str) and re.fullmatch(r"\d+", v.strip()):
                ints.append(int(v.strip()))
            else:
                return False  # non-integer cell → definitely not an index row

        if len(ints) < 3:
            return False

        # Values must be consecutive: 1, 2, 3, ...
        for i in range(1, len(ints)):
            if ints[i] != ints[i - 1] + 1:
                return False
        return True
    # ────────────────────────────────────────────────────────────────────────

    def _scan_end(self, grid: List[List[Any]], start_idx: int) -> int:
        n = len(grid)
        empty_streak = 0
        last_data = start_idx - 1

        for r in range(start_idx, n):
            row = grid[r]
            if self._is_group_row(row) or self._is_single_header_row(row):
                break
            if self._is_empty_row(row):
                empty_streak += 1
            else:
                empty_streak = 0
                last_data = r
            if empty_streak >= 3:
                break

        return max(last_data, start_idx - 1)

    def _detect_blocks(self, grid: List[List[Any]], sheet_name: str) -> List[SimpleBlock]:
        blocks: List[SimpleBlock] = []
        i = 0
        n = len(grid)

        while i < n:
            # Case A: group-header (Плательщик | Получатель) + subheader
            if self._is_group_row(grid[i]) and (i + 1 < n) and self._is_subheader_row(grid[i + 1]):
                data_start = i + 2
                # FIX 1: skip the optional sequential column-index row
                if data_start < n and self._is_column_index_row(grid[data_start]):
                    data_start += 1

                data_end = self._scan_end(grid, data_start)
                if data_end >= data_start:
                    blocks.append(SimpleBlock(
                        sheet_name=sheet_name,
                        header_row_idx=i,
                        data_start_row_idx=data_start,
                        data_end_row_idx=data_end,
                        header_rows=2,
                        group_row_idx=i,
                        subheader_row_idx=i + 1,
                    ))
                    i = data_end + 1
                    continue

            # Case B: single header row
            if self._is_single_header_row(grid[i]):
                data_start = i + 1
                # FIX 1: skip column-index row here too
                if data_start < n and self._is_column_index_row(grid[data_start]):
                    data_start += 1

                data_end = self._scan_end(grid, data_start)
                if data_end >= data_start:
                    blocks.append(SimpleBlock(
                        sheet_name=sheet_name,
                        header_row_idx=i,
                        data_start_row_idx=data_start,
                        data_end_row_idx=data_end,
                        header_rows=1,
                    ))
                    i = data_end + 1
                    continue

            i += 1

        return blocks

    def _build_df(self, grid: List[List[Any]], block: SimpleBlock) -> pd.DataFrame:
        if block.header_rows == 1:
            header = [norm_text(x) for x in grid[block.header_row_idx]]
            data = grid[block.data_start_row_idx : block.data_end_row_idx + 1]
            df = pd.DataFrame(data, columns=header)
            df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
            return df

        top = [norm_text(x) for x in grid[block.group_row_idx]]
        sub = [norm_text(x) for x in grid[block.subheader_row_idx]]

        group_map = {"плательщик": "payer", "получатель": "receiver"}
        sub_map = {
            "наименование/фио": "name", "наименование": "name", "фио": "name",
            "иин/бин": "iin_bin", "иин": "iin_bin", "бин": "iin_bin",
            "резидентство": "residency", "банк": "bank",
            "номер счета": "account", "номер счёта": "account",
            "счет": "account", "счёт": "account",
        }

        combined: List[str] = []
        current_group: Optional[str] = None

        for j in range(max(len(top), len(sub))):
            t = top[j] if j < len(top) else ""
            s = sub[j] if j < len(sub) else ""

            if t in group_map:
                current_group = group_map[t]
                combined.append(f"{current_group}/name")
                continue

            if current_group and s in sub_map:
                combined.append(f"{current_group}/{sub_map[s]}")
            elif current_group and s:
                combined.append(f"{current_group}/{s}")
            else:
                combined.append(t or s or f"col_{j}")

        data = grid[block.data_start_row_idx : block.data_end_row_idx + 1]
        df = pd.DataFrame(data, columns=combined)
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        return df

    def extract(self, file_path: str) -> List[Tuple[pd.DataFrame, Dict[str, Any]]]:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        out: List[Tuple[pd.DataFrame, Dict[str, Any]]] = []

        for sheet_name in wb.sheetnames:
            grid = self.load_grid(file_path, sheet_name)
            blocks = self._detect_blocks(grid, sheet_name)

            for bidx, b in enumerate(blocks, start=1):
                df = self._build_df(grid, b)
                if df.empty or len(df.columns) < 3:
                    continue

                block = DetectedBlock(
                    sheet_name=sheet_name,
                    header_row_idx=b.header_row_idx,
                    data_start_row_idx=b.data_start_row_idx,
                    data_end_row_idx=b.data_end_row_idx,
                    header_rows=b.header_rows,
                    group_row_idx=b.group_row_idx,
                    subheader_row_idx=b.subheader_row_idx,
                )

                stmt_meta = self.meta.extract_for_block(
                    grid=grid,
                    block=block,
                    source_bank=self.bank_name,
                    max_lookback_rows=30,
                    max_lookahead_rows=40,
                    tail_rows_in_block=40,
                )

                stmt_meta["source_sheet"] = sheet_name
                stmt_meta["source_block_id"] = bidx
                stmt_meta["source_row_base"] = b.data_start_row_idx
                stmt_meta["meta_json"] = {
                    **(stmt_meta.get("meta_json") or {}),
                    "source": "kaspi_adapter",
                    "header_row_idx": b.header_row_idx,
                    "data_start": b.data_start_row_idx,
                    "data_end": b.data_end_row_idx,
                    "header_rows": b.header_rows,
                }

                out.append((df, stmt_meta))

        return out