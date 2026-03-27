#!/usr/bin/env python3
# scripts/debug_halyk.py
from __future__ import annotations

import argparse
import os
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.utils.text_utils import norm_text
from app.ingestion.extractor.dataframe_cleaner import clean_dataframe
from app.ingestion.validation.validators import is_service_row
from app.ingestion.metadata.statement_meta_extractor import StatementMetadataExtractor


# --- Halyk multi-table header markers (same idea as adapter) ---
HALYK_HEADER_MARKERS = [
    "дата и время операции",
    "дата операции",
    "дата/время",
]


def load_grid(path: str, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    grid: List[List[Any]] = []
    max_col = ws.max_column
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row[:max_col]))
    return grid


def is_header_row(row: List[Any]) -> bool:
    joined = " ".join(norm_text(x) for x in row if x is not None)
    if not joined:
        return False
    return any(m in joined for m in HALYK_HEADER_MARKERS)


def scan_end(grid: List[List[Any]], start_idx: int) -> int:
    n = len(grid)
    empty_streak = 0
    last_data = start_idx - 1

    for r in range(start_idx, n):
        row = grid[r]

        # next table starts
        if is_header_row(row):
            break

        tokens = [norm_text(c) for c in row]
        if all(t == "" for t in tokens):
            empty_streak += 1
        else:
            empty_streak = 0
            last_data = r

        if empty_streak >= 3:
            break

    return max(last_data, start_idx - 1)


def detect_blocks(grid: List[List[Any]]) -> List[Dict[str, int]]:
    blocks: List[Dict[str, int]] = []
    i = 0
    n = len(grid)
    while i < n:
        if is_header_row(grid[i]):
            header_i = i
            data_start = i + 1
            end = scan_end(grid, data_start)
            if end >= data_start:
                blocks.append(
                    {
                        "header_row_idx": header_i,
                        "data_start_row_idx": data_start,
                        "data_end_row_idx": end,
                    }
                )
            i = end + 1
        else:
            i += 1
    return blocks


def build_df(grid: List[List[Any]], block: Dict[str, int]) -> pd.DataFrame:
    header = [norm_text(x) for x in grid[block["header_row_idx"]]]
    data = grid[block["data_start_row_idx"] : block["data_end_row_idx"] + 1]
    df = pd.DataFrame(data, columns=header)
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")
    return df


def count_tx_like_rows(df: pd.DataFrame) -> Tuple[int, int]:
    """
    Rough tx count:
      - cleans DF already
      - skips service rows (итого/остаток/etc)
    Returns: (tx_rows, skipped_rows)
    """
    tx = 0
    skipped = 0
    for _, row in df.iterrows():
        vals = [row.get(c) for c in df.columns]
        if is_service_row(vals):
            skipped += 1
            continue
        tx += 1
    return tx, skipped


def list_xlsx_from_data_root(data_root: str) -> List[str]:
    # expects data_root/halyk/*.xlsx
    halyk_dir = os.path.join(data_root, "halyk")
    if not os.path.isdir(halyk_dir):
        return []
    out = []
    for f in os.listdir(halyk_dir):
        if f.lower().endswith(".xlsx"):
            out.append(os.path.join(halyk_dir, f))
    return sorted(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default=None, help="Data root folder (expects data/halyk/*.xlsx)")
    ap.add_argument("--file", action="append", default=None, help="Explicit xlsx path. Can repeat.")
    ap.add_argument("--max_tables", type=int, default=999, help="Limit tables per file in output")
    ap.add_argument("--show_meta", action="store_true", help="Print extracted statement meta (client/iban/period/etc)")
    ap.add_argument("--show_headers", action="store_true", help="Print header preview (first 25 cols)")
    args = ap.parse_args()

    files: List[str] = []
    if args.file:
        files.extend(args.file)
    if args.data:
        files.extend(list_xlsx_from_data_root(args.data))

    files = [f for f in files if os.path.exists(f)]
    if not files:
        raise SystemExit("No .xlsx files found. Use --data data OR --file path.xlsx")

    meta_extractor = StatementMetadataExtractor()

    print("=" * 100)
    print("HALYK DEBUG REPORT")
    print("=" * 100)

    grand_tables = 0
    grand_tx = 0

    for path in files:
        print("\n" + "-" * 100)
        print(f"FILE: {os.path.basename(path)}")
        print(f"PATH: {path}")

        wb = load_workbook(path, data_only=True, read_only=True)
        sheetnames = wb.sheetnames
        print(f"SHEETS: {len(sheetnames)} -> {sheetnames}")

        file_tables = 0
        file_tx = 0

        for sheet_name in sheetnames:
            grid = load_grid(path, sheet_name)
            blocks = detect_blocks(grid)

            if not blocks:
                continue

            print(f"\n  SHEET: {sheet_name}")
            print(f"    detected_tables: {len(blocks)}")

            for t_idx, block in enumerate(blocks, start=1):
                if t_idx > args.max_tables:
                    print(f"    ... (max_tables={args.max_tables} reached)")
                    break

                df = build_df(grid, block)
                df = clean_dataframe(df)

                if df.empty or len(df.columns) < 3:
                    print(f"    TABLE #{t_idx}: header@{block['header_row_idx']} -> EMPTY/SMALL (skipped)")
                    continue

                # fake block object for StatementMetadataExtractor (needs header_row_idx)
                class _B:
                    header_row_idx = block["header_row_idx"]

                stmt_meta: Dict[str, Any] = meta_extractor.extract_for_block(
                    grid=grid,
                    block=_B(),  # type: ignore[arg-type]
                    source_bank="halyk",
                    max_lookback_rows=80,
                )

                tx_rows, skipped_rows = count_tx_like_rows(df)

                file_tables += 1
                file_tx += tx_rows

                print(
                    f"    TABLE #{t_idx}: "
                    f"header_row={block['header_row_idx']} "
                    f"data_rows={len(df)} tx_like={tx_rows} skipped_service={skipped_rows} "
                    f"cols={len(df.columns)}"
                )

                if args.show_headers:
                    cols_preview = list(df.columns)[:25]
                    print(f"      headers_preview(<=25): {cols_preview}")

                if args.show_meta:
                    keep = {
                        "client_name": stmt_meta.get("client_name"),
                        "client_iin_bin": stmt_meta.get("client_iin_bin"),
                        "account_iban": stmt_meta.get("account_iban"),
                        "currency": stmt_meta.get("currency"),
                        "statement_date": stmt_meta.get("statement_date"),
                        "period_from": stmt_meta.get("period_from"),
                        "period_to": stmt_meta.get("period_to"),
                        "opening_balance": stmt_meta.get("opening_balance"),
                        "closing_balance": stmt_meta.get("closing_balance"),
                        "total_debit": stmt_meta.get("total_debit"),
                        "total_credit": stmt_meta.get("total_credit"),
                    }
                    print(f"      meta: {keep}")

        grand_tables += file_tables
        grand_tx += file_tx

        print(f"\n  FILE SUMMARY: tables={file_tables} tx_like_total={file_tx}")

    print("\n" + "=" * 100)
    print(f"GRAND SUMMARY: files={len(files)} tables={grand_tables} tx_like_total={grand_tx}")
    print("=" * 100)


if __name__ == "__main__":
    main()